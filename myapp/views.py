from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.db.models import Q, Avg, Count, Sum, Max, Min
from django.utils import timezone
from .forms import ExamForm, QuestionForm, ChoiceForm, QuestionFormSet, ChoiceFormSet, ProfileUpdateForm, TeacherCreationForm, StudentCreationForm, BulkStudentUploadForm
from .models import User, Exam, Question, Choice, ExamAttempt, StudentAnswer, StudentClass, School


# Home/Landing Page
def home(request):
    """Landing page with role-based redirect for authenticated users"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'home.html')


# Role-Based Dashboard
@login_required
def dashboard(request):
    """
    Role-based dashboard redirect:
    - Teachers: Content creators and performance analysts
    - Students: Active learners and self-evaluators
    - School Admins: User/class management within their school
    """
    if request.user.is_teacher:
        return redirect('teacher_dashboard')
    elif request.user.is_student:
        return redirect('student_dashboard')
    elif request.user.is_school_admin:
        return redirect('school_admin_dashboard')
    elif request.user.is_superuser:
        return redirect('/admin/')
    else:
        messages.error(request, 'Your account role is not configured properly. Please contact admin.')
        return redirect('home')


# Teacher Dashboard
@login_required
def teacher_dashboard(request):
    """
    Teacher Dashboard: Content creation, exam management, performance analytics
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    # Get all exams created by this teacher
    teacher_exams = Exam.objects.filter(created_by=request.user)
    assigned_exams = teacher_exams.filter(assigned_classes__isnull=False).distinct().count()
    unassigned_exams = teacher_exams.filter(assigned_classes__isnull=True).count()
    
    context = {
        'user': request.user,
        'page_title': 'Teacher Dashboard',
        'assigned_exams': assigned_exams,
        'unassigned_exams': unassigned_exams,
    }
    return render(request, 'dashboard/teacher_dashboard.html', context)


# Student Dashboard
@login_required
def student_dashboard(request):
    """
    Student Dashboard: Available exams, attempt history, performance tracking
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Get all completed attempts by this student
    completed_attempts = ExamAttempt.objects.filter(
        student=request.user,
        is_completed=True
    )
    
    # Count unique exams taken
    exams_taken = completed_attempts.values('exam').distinct().count()
    
    # Get available exams (published exams in enrolled classes that haven't been taken yet)
    enrolled_classes = StudentClass.objects.filter(
        students=request.user,
        is_active=True
    )
    
    # Get all exams assigned to student's classes
    available_exams = Exam.objects.filter(
        assigned_classes__in=enrolled_classes
    ).distinct()
    
    # Filter out exams already completed
    completed_exam_ids = completed_attempts.values_list('exam_id', flat=True).distinct()
    pending_exams = available_exams.exclude(id__in=completed_exam_ids).count()
    
    # Calculate leaderboard ranking across all enrolled classes
    student_ranking = None
    total_students_in_classes = 0
    
    if enrolled_classes.exists():
        # Get the first enrolled class for ranking
        first_class = enrolled_classes.first()
        if first_class:
            # CRITICAL: Only use exams CURRENTLY assigned to this class
            # This automatically excludes any exams that were unassigned from the class
            # All exam types included, but only Test mode attempts count towards leaderboard
            assigned_exams = first_class.assigned_exams.all()
            assigned_exam_ids = list(assigned_exams.values_list('id', flat=True))
            
            if assigned_exams.exists():
                # Get all students in class (including current user, regardless of is_active)
                students = first_class.students.filter(role='student')
                
                if students.exists():
                    # Calculate average percentage for all students (consistent with leaderboard logic)
                    student_scores = []
                    for student in students:
                        total_percentage_sum = 0
                        exams_attempted = 0
                        total_time_seconds = 0
                        
                        for exam in assigned_exams:
                            if exam.total_marks == 0:
                                continue
                            # Get BEST TEST attempt (highest score) for this exam
                            # CRITICAL: Only from exams currently assigned to this class
                            # Only test mode attempts count for leaderboard
                            best_attempt = ExamAttempt.objects.filter(
                                student=student,
                                exam=exam,
                                exam_id__in=assigned_exam_ids,  # Explicit filter: only currently assigned exams
                                student_class=first_class,
                                is_completed=True,
                                attempt_mode='test'  # Only test attempts count
                            ).order_by('-score').first()
                            
                            if best_attempt:
                                # Calculate percentage for this exam (cap at 100%)
                                exam_percentage = (float(best_attempt.score) / float(exam.total_marks) * 100) if exam.total_marks > 0 else 0
                                exam_percentage = min(exam_percentage, 100.0)
                                
                                # Calculate time in seconds
                                if best_attempt.submitted_at:
                                    time_taken_seconds = (best_attempt.submitted_at - best_attempt.started_at).total_seconds()
                                else:
                                    time_taken_seconds = 0
                                
                                total_percentage_sum += exam_percentage
                                total_time_seconds += time_taken_seconds
                                exams_attempted += 1
                        
                        # Only include students with at least one completed attempt
                        if exams_attempted > 0:
                            average_percentage = total_percentage_sum / exams_attempted
                            average_time_seconds = total_time_seconds / exams_attempted
                            student_scores.append({
                                'student_id': student.id,
                                'exams_attempted': exams_attempted,
                                'average_percentage': average_percentage,
                                'average_time_seconds': average_time_seconds,
                            })
                    
                    if student_scores:
                        # Sort by ranking priority (consistent with class_leaderboard):
                        # 1. Number of Test exams taken (more tests = higher rank)
                        # 2. Average Test score percentage (higher percentage = higher rank)
                        # 3. Average Test completion time (lower time = higher rank)
                        student_scores.sort(key=lambda x: (-x['exams_attempted'], -x['average_percentage'], x['average_time_seconds']))
                        
                        # Assign ranks (handle ties: same on all three criteria = same rank)
                        for i, data in enumerate(student_scores):
                            if i > 0:
                                prev_data = student_scores[i - 1]
                                if (data['exams_attempted'] == prev_data['exams_attempted'] and
                                    data['average_percentage'] == prev_data['average_percentage'] and
                                    data['average_time_seconds'] == prev_data['average_time_seconds']):
                                    data['rank'] = prev_data['rank']
                                else:
                                    data['rank'] = i + 1
                            else:
                                data['rank'] = 1
                        
                        # Find current student's rank
                        total_students_in_classes = len(student_scores)
                        for data in student_scores:
                            if data['student_id'] == request.user.id:
                                student_ranking = data['rank']
                                break
    
    context = {
        'user': request.user,
        'page_title': 'Student Dashboard',
        'exams_taken': exams_taken,
        'available_exams': pending_exams,
        'student_ranking': student_ranking,
        'total_students_in_classes': total_students_in_classes,
    }
    return render(request, 'dashboard/student_dashboard.html', context)


# Signup disabled — public registration is not allowed
def signup_disabled(request):
    """Public self-registration is disabled. Accounts are created by school admins."""
    messages.info(
        request,
        'Self-registration is disabled. '
        'Teacher and student accounts are created by your school administrator. '
        'Please contact your school admin to get access.'
    )
    return redirect('account_login')

from django.contrib.auth import login
from .forms import SchoolRegistrationForm

def register_school(request):
    """Public view to register a new school and its first admin"""
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        form = SchoolRegistrationForm(request.POST)
        if form.is_valid():
            school_name = form.cleaned_data['school_name']
            
            # Generate a strict 6-character unique alphanumeric school code
            import random
            import string
            
            def get_random_code():
                while True:
                    # Generate 6 char uppercase alphanumeric
                    chars = random.choices(string.ascii_uppercase + string.digits, k=6)
                    code = ''.join(chars)
                    # Enforce that it contains at least one letter and at least one digit
                    if any(c.isdigit() for c in code) and any(c.isalpha() for c in code):
                        if not School.objects.filter(code=code).exists():
                            return code

            code = get_random_code()
            
            # Create School
            school = School.objects.create(
                name=school_name,
                code=code
            )
            
            # Create Admin User
            # username format: <school_code>_ADMIN
            username = f"{code}_ADMIN"
            admin_user = User.objects.create_user(
                username=username,
                email=form.cleaned_data['admin_email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['admin_first_name'],
                last_name=form.cleaned_data['admin_last_name'],
                phone_number=form.cleaned_data['admin_phone'],
                role='admin',
                school=school,
                is_staff=False,
                is_superuser=False
            )
            
            messages.success(request, f'School registered successfully! Welcome {admin_user.first_name}.')
            login(request, admin_user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect('dashboard')
    else:
        form = SchoolRegistrationForm()
        
    return render(request, 'account/register_school.html', {'form': form, 'page_title': 'Register Your School'})



# ============================================================================
# EXAM MANAGEMENT VIEWS (Teacher Only)
# ============================================================================

class TeacherRequiredMixin(UserPassesTestMixin):
    """Mixin to ensure only teachers can access certain views"""
    
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_teacher
    
    def handle_no_permission(self):
        messages.error(self.request, 'Access denied. Teachers only.')
        return redirect('dashboard')


class ExamListView(LoginRequiredMixin, TeacherRequiredMixin, ListView):
    """
    List all exams created by the logged-in teacher (school-scoped)
    """
    model = Exam
    template_name = 'exams/exam_list.html'
    context_object_name = 'exams'
    paginate_by = None

    def get_queryset(self):
        # School-scoped: only exams belonging to the teacher's school
        return Exam.objects.filter(
            created_by=self.request.user,
            school=self.request.user.school
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'My Exams'
        all_exams = self.get_queryset()
        context['total_exams'] = all_exams.count()
        context['assigned_exams'] = all_exams.filter(assigned_classes__isnull=False).distinct().count()
        from collections import defaultdict
        exams_by_subject = defaultdict(list)
        for exam in all_exams:
            exams_by_subject[exam.subject].append(exam)
        context['exams_by_subject'] = sorted(exams_by_subject.items(), key=lambda x: x[0])
        return context


class ExamDetailView(LoginRequiredMixin, TeacherRequiredMixin, DetailView):
    """View exam details — school-scoped"""
    model = Exam
    template_name = 'exams/exam_detail.html'
    context_object_name = 'exam'

    def get_queryset(self):
        return Exam.objects.filter(
            created_by=self.request.user,
            school=self.request.user.school
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['questions'] = self.object.questions.prefetch_related('choices').all()
        context['page_title'] = f'Exam: {self.object.title}'
        return context


class ExamCreateView(LoginRequiredMixin, TeacherRequiredMixin, CreateView):
    """Create a new exam — auto-assigns school from teacher."""
    model = Exam
    form_class = ExamForm
    template_name = 'exams/exam_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.school = self.request.user.school  # school-scoped
        messages.success(self.request, f'Exam "{form.instance.title}" created successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('exam_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Exam'
        context['submit_text'] = 'Create Exam'
        return context


class ExamUpdateView(LoginRequiredMixin, TeacherRequiredMixin, UpdateView):
    """Update an existing exam — school-scoped"""
    model = Exam
    form_class = ExamForm
    template_name = 'exams/exam_form.html'

    def get_queryset(self):
        return Exam.objects.filter(
            created_by=self.request.user,
            school=self.request.user.school
        )

    def form_valid(self, form):
        messages.success(self.request, f'Exam "{form.instance.title}" updated successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('exam_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit: {self.object.title}'
        context['submit_text'] = 'Update Exam'
        return context


class ExamDeleteView(LoginRequiredMixin, TeacherRequiredMixin, DeleteView):
    """Delete an exam — school-scoped"""
    model = Exam
    template_name = 'exams/exam_confirm_delete.html'
    success_url = reverse_lazy('exam_list')

    def get_queryset(self):
        return Exam.objects.filter(
            created_by=self.request.user,
            school=self.request.user.school
        )

    def delete(self, request, *args, **kwargs):
        exam = self.get_object()
        messages.success(request, f'Exam "{exam.title}" deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# QUESTION MANAGEMENT VIEWS
# ============================================================================

@login_required
def question_edit(request, exam_pk, question_pk):
    """Edit an existing question"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    question = get_object_or_404(Question, pk=question_pk, exam=exam)
    
    if request.method == 'POST':
        question_form = QuestionForm(request.POST, instance=question)
        choice_formset = ChoiceFormSet(request.POST, instance=question)
        
        if question_form.is_valid() and choice_formset.is_valid():
            # Get the choices that will remain after save (excluding deleted ones)
            remaining_choices = []
            correct_count = 0
            
            for form in choice_formset:
                # Skip forms marked for deletion
                if form.cleaned_data.get('DELETE', False):
                    continue
                
                # Skip empty forms
                if not form.cleaned_data.get('choice_text', '').strip():
                    continue
                
                remaining_choices.append(form.cleaned_data)
                if form.cleaned_data.get('is_correct', False):
                    correct_count += 1
            
            # Validate we have exactly 4 choices
            if len(remaining_choices) != 4:
                messages.error(request, f'Each question must have exactly 4 choices. You have {len(remaining_choices)} choices.')
                context = {
                    'exam': exam,
                    'question': question,
                    'question_form': question_form,
                    'choice_formset': choice_formset,
                    'page_title': f'Edit Question'
                }
                return render(request, 'exams/question_edit.html', context)
            
            # Validate exactly one correct answer
            if correct_count != 1:
                if correct_count == 0:
                    messages.error(request, 'Please mark exactly ONE choice as correct. No choice is currently marked.')
                else:
                    messages.error(request, f'Please mark exactly ONE choice as correct. You have {correct_count} choices marked.')
                context = {
                    'exam': exam,
                    'question': question,
                    'question_form': question_form,
                    'choice_formset': choice_formset,
                    'page_title': f'Edit Question'
                }
                return render(request, 'exams/question_edit.html', context)
            
            # Validate all choices have text
            for i, choice_data in enumerate(remaining_choices):
                choice_text = choice_data.get('choice_text', '').strip()
                if len(choice_text) < 1:
                    messages.error(request, f'Choice {chr(65+i)} cannot be empty.')
                    context = {
                        'exam': exam,
                        'question': question,
                        'question_form': question_form,
                        'choice_formset': choice_formset,
                        'page_title': f'Edit Question'
                    }
                    return render(request, 'exams/question_edit.html', context)
            
            # All validations passed - save the data
            try:
                with transaction.atomic():
                    # Save the question
                    updated_question = question_form.save(commit=False)
                    updated_question.save()
                    
                    # Save the choices and ensure order integrity
                    saved_choices = choice_formset.save(commit=False)
                    
                    # Ensure each choice has a proper order value (0, 1, 2, 3)
                    for idx, choice in enumerate(saved_choices):
                        if choice.order is None or choice.order < 0:
                            choice.order = idx
                        choice.save()
                    
                    # Handle deletions (if any were marked)
                    for obj in choice_formset.deleted_objects:
                        obj.delete()
                    
                    # Verify we still have exactly 4 choices after save
                    final_choice_count = question.choices.count()
                    if final_choice_count != 4:
                        raise ValueError(f'Expected 4 choices after save, but got {final_choice_count}')
                    
                    # Update total marks for the exam
                    exam.update_total_marks()
                    
                    messages.success(request, 'Question updated successfully!')
                    return redirect('exam_detail', pk=exam.pk)
            except Exception as e:
                messages.error(request, f'Error saving question: {str(e)}')
        else:
            # Form validation failed
            if question_form.errors:
                for field, errors in question_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            
            if choice_formset.errors:
                for i, form_errors in enumerate(choice_formset.errors):
                    if form_errors:
                        for field, errors in form_errors.items():
                            for error in errors:
                                messages.error(request, f'Choice {chr(65+i)} - {field}: {error}')
            
            if choice_formset.non_form_errors():
                for error in choice_formset.non_form_errors():
                    messages.error(request, error)
    else:
        question_form = QuestionForm(instance=question)
        choice_formset = ChoiceFormSet(instance=question)
    
    context = {
        'exam': exam,
        'question': question,
        'question_form': question_form,
        'choice_formset': choice_formset,
        'page_title': f'Edit Question'
    }
    return render(request, 'exams/question_edit.html', context)


@login_required
def question_delete(request, exam_pk, question_pk):
    """Delete a question"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    question = get_object_or_404(Question, pk=question_pk, exam=exam)
    
    if request.method == 'POST':
        question.delete()
        # Update total marks
        exam.update_total_marks()
        messages.success(request, 'Question deleted successfully!')
        return redirect('exam_detail', pk=exam.pk)
    
    context = {
        'exam': exam,
        'question': question,
        'page_title': 'Delete Question'
    }
    return render(request, 'exams/question_confirm_delete.html', context)


@login_required
def exam_add_questions(request, exam_pk):
    """
    Add multiple questions with choices to an exam
    
    Philosophy: Efficient content creation without sacrificing quality.
    Teachers can add multiple questions with all 4 choices on a single page.
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    if request.method == 'POST':
        # Prevent duplicate form submissions (double-click, browser back, refresh)
        submission_token = request.POST.get('submission_token', '')
        session_key = f'exam_{exam_pk}_submission_{submission_token}'
        
        if submission_token and request.session.get(session_key):
            # This form was already processed
            messages.info(request, 'These questions have already been added. Redirecting to exam details.')
            return redirect('exam_detail', pk=exam.pk)
        
        # Get the number of questions submitted
        num_questions = int(request.POST.get('num_questions', 1))
        
        with transaction.atomic():
            # Get current question count BEFORE the loop to fix order bug
            base_order = exam.questions.count()
            created_count = 0
            
            for i in range(num_questions):
                # Get question data
                question_text = request.POST.get(f'question_{i}_text', '').strip()
                question_marks = request.POST.get(f'question_{i}_marks', '1')
                question_explanation = request.POST.get(f'question_{i}_explanation', '').strip()
                
                if not question_text:
                    continue
                
                # Create question with correct order
                question = Question.objects.create(
                    exam=exam,
                    question_text=question_text,
                    marks=int(question_marks),
                    explanation=question_explanation if question_explanation else None,
                    order=base_order + created_count
                )
                
                # Get choices data
                choices_data = []
                correct_count = 0
                for j in range(4):
                    choice_text = request.POST.get(f'question_{i}_choice_{j}_text', '').strip()
                    is_correct = request.POST.get(f'question_{i}_choice_{j}_correct') == 'on'
                    
                    if choice_text:
                        choices_data.append({
                            'text': choice_text,
                            'is_correct': is_correct,
                            'order': j
                        })
                        if is_correct:
                            correct_count += 1
                
                # Validate choices
                if len(choices_data) != 4:
                    question.delete()
                    messages.warning(request, f'Question {i+1}: Must have exactly 4 choices.')
                    continue
                
                if correct_count != 1:
                    question.delete()
                    messages.warning(request, f'Question {i+1}: Must have exactly 1 correct answer.')
                    continue
                
                # Create choices
                for choice_data in choices_data:
                    Choice.objects.create(
                        question=question,
                        choice_text=choice_data['text'],
                        is_correct=choice_data['is_correct'],
                        order=choice_data['order']
                    )
                
                created_count += 1
            
            if created_count > 0:
                # Update total marks
                exam.update_total_marks()
                
                # Mark this submission as processed to prevent duplicates
                if submission_token:
                    request.session[session_key] = True
                    # Auto-expire after 1 hour
                    request.session.set_expiry(3600)
                
                messages.success(request, f'Successfully added {created_count} question(s) to "{exam.title}"!')
                return redirect('exam_detail', pk=exam.pk)
            else:
                messages.error(request, 'No valid questions were added. Please check your input.')
    
    context = {
        'exam': exam,
        'page_title': f'Add Questions - {exam.title}',
    }
    return render(request, 'exams/exam_add_questions.html', context)


@login_required
def download_question_template(request, exam_pk):
    """
    Generate and download Excel template for bulk question upload
    
    Philosophy: Provide teachers with a structured template to ensure data consistency
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions Template"
    
    # Define headers
    headers = [
        'Question Text',
        'Option A',
        'Option B', 
        'Option C',
        'Option D',
        'Correct Answer (A/B/C/D)',
        'Explanation (Optional)',
        'Marks'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add sample data
    sample_data = [
        [
            'What is the capital of France?',
            'London',
            'Paris',
            'Berlin',
            'Madrid',
            'B',
            'Paris is the capital and largest city of France.',
            '1'
        ],
        [
            'Which planet is known as the Red Planet?',
            'Venus',
            'Mars',
            'Jupiter',
            'Saturn',
            'B',
            'Mars appears red due to iron oxide on its surface.',
            '1'
        ],
        [
            '2 + 2 = ?',
            '3',
            '4',
            '5',
            '6',
            'B',
            '-',
            '1'
        ]
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    column_widths = [40, 25, 25, 25, 25, 20, 40, 10]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="questions_template_{exam.title.replace(" ", "_")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def bulk_upload_questions(request, exam_pk):
    """
    Upload Excel file with questions - Step 1: Parse and Preview
    
    Philosophy: Parse Excel, validate all data, show preview for teacher confirmation
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        from openpyxl import load_workbook
        
        excel_file = request.FILES['excel_file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file (.xlsx or .xls)')
            return redirect('bulk_upload_questions', exam_pk=exam.pk)
        
        try:
            # Load workbook
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
            
            parsed_questions = []
            errors = []
            row_num = 2  # Start from row 2 (skip header)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(row):
                    continue
                
                # Extract data from row
                question_text = str(row[0]).strip() if row[0] else ''
                option_a = str(row[1]).strip() if row[1] else ''
                option_b = str(row[2]).strip() if row[2] else ''
                option_c = str(row[3]).strip() if row[3] else ''
                option_d = str(row[4]).strip() if row[4] else ''
                correct_answer = str(row[5]).strip().upper() if row[5] else ''
                explanation = str(row[6]).strip() if row[6] else ''
                marks = str(row[7]).strip() if row[7] else ''
                
                # Handle explanation: treat empty or "-" as no explanation
                if not explanation or explanation == '-':
                    explanation = None
                
                # Validation
                row_errors = []
                marks_int = 1  # Default value
                
                if not question_text:
                    row_errors.append('Question text is required')
                
                if not option_a:
                    row_errors.append('Option A is required')
                if not option_b:
                    row_errors.append('Option B is required')
                if not option_c:
                    row_errors.append('Option C is required')
                if not option_d:
                    row_errors.append('Option D is required')
                
                if not correct_answer:
                    row_errors.append('Correct answer is required')
                elif correct_answer not in ['A', 'B', 'C', 'D']:
                    row_errors.append('Correct answer must be A, B, C, or D')
                
                if not marks:
                    row_errors.append('Marks field is required')
                else:
                    try:
                        marks_int = int(marks)
                        if marks_int < 1:
                            row_errors.append('Marks must be at least 1')
                    except (ValueError, TypeError):
                        row_errors.append('Marks must be a valid number')
                
                if row_errors:
                    errors.append({
                        'row': row_num,
                        'errors': row_errors
                    })
                else:
                    # Map correct answer to choice
                    correct_choice_map = {
                        'A': 0,
                        'B': 1,
                        'C': 2,
                        'D': 3
                    }
                    
                    parsed_questions.append({
                        'question_text': question_text,
                        'options': [option_a, option_b, option_c, option_d],
                        'correct_choice_index': correct_choice_map[correct_answer],
                        'correct_answer_letter': correct_answer,
                        'explanation': explanation,
                        'marks': marks_int,
                        'row_number': row_num
                    })
                
                row_num += 1
            
            wb.close()
            
            # Check if any questions were parsed
            if not parsed_questions and not errors:
                messages.error(request, 'No questions found in the Excel file. Please check the file and try again.')
                return redirect('bulk_upload_questions', exam_pk=exam.pk)
            
            # Store parsed data in session for confirmation
            request.session['bulk_upload_data'] = {
                'exam_pk': exam.pk,
                'questions': parsed_questions,
                'errors': errors,
                'total_rows': row_num - 2
            }
            
            return redirect('bulk_upload_preview', exam_pk=exam.pk)
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
            return redirect('bulk_upload_questions', exam_pk=exam.pk)
    
    context = {
        'exam': exam,
        'page_title': f'Bulk Upload Questions - {exam.title}',
    }
    return render(request, 'exams/bulk_upload_questions.html', context)


@login_required
def bulk_upload_preview(request, exam_pk):
    """
    Preview parsed questions - Step 2: Show preview and confirm
    
    Philosophy: Allow teachers to review all parsed data before committing to database
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    # Retrieve parsed data from session
    bulk_data = request.session.get('bulk_upload_data')
    
    if not bulk_data or bulk_data.get('exam_pk') != exam.pk:
        messages.error(request, 'No upload data found. Please upload the Excel file again.')
        return redirect('bulk_upload_questions', exam_pk=exam.pk)
    
    context = {
        'exam': exam,
        'parsed_questions': bulk_data.get('questions', []),
        'errors': bulk_data.get('errors', []),
        'total_rows': bulk_data.get('total_rows', 0),
        'page_title': f'Preview Bulk Upload - {exam.title}',
    }
    return render(request, 'exams/bulk_upload_preview.html', context)


@login_required
def bulk_upload_confirm(request, exam_pk):
    """
    Save all questions to database - Step 3: Atomic save operation
    
    Philosophy: All-or-nothing transaction to prevent partial imports and duplicates
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    if request.method != 'POST':
        return redirect('bulk_upload_questions', exam_pk=exam.pk)
    
    # Retrieve parsed data from session
    bulk_data = request.session.get('bulk_upload_data')
    
    if not bulk_data or bulk_data.get('exam_pk') != exam.pk:
        messages.error(request, 'No upload data found. Please upload the Excel file again.')
        return redirect('bulk_upload_questions', exam_pk=exam.pk)
    
    parsed_questions = bulk_data.get('questions', [])
    
    if not parsed_questions:
        messages.error(request, 'No valid questions to save.')
        return redirect('bulk_upload_questions', exam_pk=exam.pk)
    
    try:
        # Atomic transaction - all or nothing
        with transaction.atomic():
            base_order = exam.questions.count()
            created_count = 0
            
            for idx, q_data in enumerate(parsed_questions):
                # Create question
                question = Question.objects.create(
                    exam=exam,
                    question_text=q_data['question_text'],
                    marks=q_data['marks'],
                    explanation=q_data['explanation'],
                    order=base_order + idx
                )
                
                # Create choices
                for choice_idx, option_text in enumerate(q_data['options']):
                    Choice.objects.create(
                        question=question,
                        choice_text=option_text,
                        is_correct=(choice_idx == q_data['correct_choice_index']),
                        order=choice_idx
                    )
                
                created_count += 1
            
            # Update exam total marks
            exam.update_total_marks()
            
            # Clear session data
            if 'bulk_upload_data' in request.session:
                del request.session['bulk_upload_data']
            
            messages.success(
                request, 
                f'Successfully added {created_count} question(s) to "{exam.title}"! '
                f'Total questions: {exam.questions.count()}'
            )
            return redirect('exam_detail', pk=exam.pk)
            
    except Exception as e:
        messages.error(request, f'Error saving questions: {str(e)}. No questions were added.')
        return redirect('bulk_upload_preview', exam_pk=exam.pk)


# ============================================================================
# STUDENT EXAM TAKING VIEWS
# ============================================================================

@login_required
def student_exam_list(request):
    """
    List all available exams for students
    
    Philosophy: Students discover learning opportunities with practice-oriented retakes
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Get student's classes
    student_classes = request.user.student_classes.all()
    
    # Get all exams assigned to student's classes (no unassigned exams visible)
    available_exams = Exam.objects.filter(
        assigned_classes__in=student_classes
    ).distinct().prefetch_related('questions', 'assigned_classes')
    
    # Get student's attempts (class-scoped)
    student_attempts = ExamAttempt.objects.filter(
        student=request.user,
        is_completed=True,
        student_class__in=student_classes  # Only show attempts from active classes
    ).select_related('exam', 'student_class').order_by('exam_id', '-submitted_at')
    
    # Create dictionaries for exam data
    exam_attempts = {}  # exam_id -> list of attempts
    exam_test_attempts = {}  # exam_id -> list of test attempts
    exam_practice_attempts = {}  # exam_id -> list of practice attempts
    exam_best_score = {}  # exam_id -> best percentage (from test attempts)
    exam_latest_attempt = {}  # exam_id -> latest attempt
    exam_has_test_attempt = {}  # exam_id -> boolean (has taken test attempt)
    
    for attempt in student_attempts:
        exam_id = attempt.exam_id
        if exam_id not in exam_attempts:
            exam_attempts[exam_id] = []
        exam_attempts[exam_id].append(attempt)
        
        # Track test vs practice attempts
        if attempt.attempt_mode == 'test':
            if exam_id not in exam_test_attempts:
                exam_test_attempts[exam_id] = []
            exam_test_attempts[exam_id].append(attempt)
            exam_has_test_attempt[exam_id] = True
            
            # Track best score from test attempts only
            if exam_id not in exam_best_score or attempt.percentage > exam_best_score[exam_id]:
                exam_best_score[exam_id] = attempt.percentage
        else:  # practice mode
            if exam_id not in exam_practice_attempts:
                exam_practice_attempts[exam_id] = []
            exam_practice_attempts[exam_id].append(attempt)
        
        # Track latest attempt (already ordered by -submitted_at)
        if exam_id not in exam_latest_attempt:
            exam_latest_attempt[exam_id] = attempt
    
    # Annotate exams with attempt info
    for exam in available_exams:
        exam.attempt_list = exam_attempts.get(exam.id, [])
        exam.attempt_count = len(exam.attempt_list)
        exam.test_attempts = exam_test_attempts.get(exam.id, [])
        exam.practice_attempts = exam_practice_attempts.get(exam.id, [])
        exam.has_test_attempt = exam_has_test_attempt.get(exam.id, False)
        exam.best_score = exam_best_score.get(exam.id)
        exam.latest_attempt = exam_latest_attempt.get(exam.id)
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        available_exams = available_exams.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    # Group exams by subject for organized display
    from collections import defaultdict
    exams_by_subject = defaultdict(list)
    for exam in available_exams:
        exams_by_subject[exam.subject].append(exam)
    
    # Sort subjects alphabetically and convert to list of tuples
    exams_by_subject = sorted(exams_by_subject.items(), key=lambda x: x[0])
    
    # Calculate statistics for scalability
    total_exam_count = sum(len(exams) for _, exams in exams_by_subject)
    subject_stats = {subject: len(exams) for subject, exams in exams_by_subject}
    
    context = {
        'available_exams': available_exams,
        'exams_by_subject': exams_by_subject,  # Grouped exams for subject-wise display
        'student_classes': student_classes,
        'page_title': 'Available Exams',
        'search_query': search_query,
        'total_exam_count': total_exam_count,
        'subject_stats': subject_stats,
    }
    return render(request, 'exams/student_exam_list.html', context)


@login_required
def exam_take(request, exam_pk):
    """
    Take an exam (displays questions with timer)
    
    Philosophy: Time-bound assessment with clear UX. Students can retake for practice.
    Class-scoped: Each attempt is linked to a specific class context.
    
    Question Randomization (Per-Attempt):
    - Questions are shuffled uniquely for each exam attempt
    - Question order randomization enhances assessment integrity
    - Only presentation order changes; question IDs remain unchanged
    - Form submission uses question.id (order-independent)
    - Score calculation uses question.id (order-independent)
    - Results display in original exam order (by 'order' field)
    - Teacher views display in original exam order
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Verify exam exists and student has access via their class
    student_classes = request.user.student_classes.all()
    exam = get_object_or_404(Exam, pk=exam_pk, assigned_classes__in=student_classes)
    
    # Get class context from query parameter or determine from student's classes
    class_id = request.GET.get('class_id')
    if class_id:
        student_class = get_object_or_404(
            StudentClass,
            pk=class_id,
            students=request.user,
            is_active=True
        )
    else:
        # Auto-select first available class if student is in only one class with this exam
        student_classes = request.user.student_classes.filter(
            is_active=True,
            assigned_exams=exam
        )
        if student_classes.count() == 1:
            student_class = student_classes.first()
        elif student_classes.count() == 0:
            # Check if exam is not assigned to any class (legacy)
            if exam.assigned_classes.count() == 0:
                # Use any of student's active classes
                student_class = request.user.student_classes.filter(is_active=True).first()
                if not student_class:
                    messages.error(request, 'You must be enrolled in at least one class to take exams.')
                    return redirect('student_exam_list')
            else:
                messages.error(request, 'You do not have access to this exam.')
                return redirect('student_exam_list')
        else:
            # Multiple classes - need to select
            messages.info(request, 'Please select a class to take this exam in.')
            return redirect('student_exam_list')
    
    # Check previous attempts (class-scoped)
    previous_attempts = ExamAttempt.objects.filter(
        student=request.user,
        exam=exam,
        student_class=student_class,
        is_completed=True
    ).order_by('-submitted_at')
    
    # Get attempt mode from query parameter (for practice_test type)
    attempt_mode = request.GET.get('mode', 'test')  # Default to 'test' mode
    
    # Validate attempt mode based on exam type
    if exam.exam_type == 'test':
        # For pure 'test' exams, always use test mode
        attempt_mode = 'test'
        # Enforce single-attempt restriction
        if previous_attempts.exists():
            messages.error(
                request, 
                f'You have already completed this test exam. Test exams allow only one attempt. '
                f'Your score: {previous_attempts.first().score}/{previous_attempts.first().total_marks}'
            )
            return redirect('exam_results', attempt_pk=previous_attempts.first().pk)
    elif exam.exam_type == 'practice_test':
        # For practice_test exams, check mode-specific restrictions
        if attempt_mode == 'test':
            # Check if student has already taken a test attempt
            test_attempts = previous_attempts.filter(attempt_mode='test')
            if test_attempts.exists():
                messages.error(
                    request, 
                    f'You have already completed the Test attempt for this exam. You can only take one Test attempt. '
                    f'Your Test score: {test_attempts.first().score}/{test_attempts.first().total_marks}. '
                    f'You may continue taking Practice attempts for learning.'
                )
                return redirect('exam_results', attempt_pk=test_attempts.first().pk)
        # Practice mode has no restrictions - unlimited attempts allowed
    
    # Create new attempt with specified mode
    attempt = ExamAttempt.objects.create(
        student=request.user,
        exam=exam,
        student_class=student_class,
        total_marks=exam.total_marks,
        attempt_mode=attempt_mode
    )
    
    # Get questions with choices
    questions = list(exam.questions.prefetch_related('choices').order_by('order'))
    
    # Randomize question order for each attempt (per-attempt randomization)
    import random
    random.shuffle(questions)
    
    # Save the randomized question order to the attempt
    attempt.question_order = [q.id for q in questions]
    attempt.save(update_fields=['question_order'])
    
    # Randomize choice order for each question (per-attempt randomization)
    questions_with_shuffled_choices = []
    for question in questions:
        # Convert to list and shuffle
        choices_list = list(question.choices.all())
        random.shuffle(choices_list)
        # Attach shuffled choices back to question object
        question.shuffled_choices = choices_list
        questions_with_shuffled_choices.append(question)
    
    context = {
        'exam': exam,
        'attempt': attempt,
        'questions': questions_with_shuffled_choices,
        'questions_count': len(questions_with_shuffled_choices),
        'student_class': student_class,
        'previous_attempts': previous_attempts,
        'attempt_number': previous_attempts.count() + 1,
        'page_title': f'Taking: {exam.title}'
    }
    return render(request, 'exams/exam_take.html', context)


@login_required
def exam_submit(request, attempt_pk):
    """
    Submit exam and calculate score
    
    Philosophy: Process answers and provide immediate learning feedback
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    attempt = get_object_or_404(ExamAttempt, pk=attempt_pk, student=request.user)
    
    if attempt.is_completed:
        messages.info(request, 'This exam has already been submitted.')
        return redirect('exam_results', attempt_pk=attempt.pk)
    
    # Check if exam time has ended
    from datetime import timedelta
    end_time = attempt.started_at + timedelta(minutes=attempt.exam.duration_minutes)
    if timezone.now() > end_time:
        messages.error(request, 'Time is up! The exam has ended and cannot be submitted.')
        # Auto-submit with current answers
        with transaction.atomic():
            questions = attempt.exam.questions.all()
            for question in questions:
                choice_id = request.POST.get(f'question_{question.id}') if request.method == 'POST' else None
                if choice_id:
                    try:
                        selected_choice = Choice.objects.get(id=choice_id, question=question)
                        StudentAnswer.objects.create(
                            attempt=attempt,
                            question=question,
                            selected_choice=selected_choice
                        )
                    except Choice.DoesNotExist:
                        pass
            attempt.submitted_at = timezone.now()
            attempt.is_completed = True
            attempt.calculate_score()
            attempt.save()
        return redirect('exam_results', attempt_pk=attempt.pk)
    
    if request.method == 'POST':
        with transaction.atomic():
            # Check for violation data from auto-submit
            has_violation = request.POST.get('has_violation') == 'true'
            violation_reason = request.POST.get('violation_reason', '')
            
            # Process each answer
            questions = attempt.exam.questions.all()
            
            for question in questions:
                choice_id = request.POST.get(f'question_{question.id}')
                
                if choice_id:
                    try:
                        selected_choice = Choice.objects.get(
                            id=choice_id,
                            question=question
                        )
                        
                        # Create student answer
                        StudentAnswer.objects.create(
                            attempt=attempt,
                            question=question,
                            selected_choice=selected_choice
                        )
                    except Choice.DoesNotExist:
                        pass
            
            # Mark as completed and calculate score
            attempt.submitted_at = timezone.now()
            attempt.is_completed = True
            
            # Record violation if detected
            if has_violation:
                attempt.has_violation = True
                attempt.violation_reason = violation_reason
            
            attempt.calculate_score()
            attempt.save()
            
            messages.success(request, f'Exam submitted successfully! You scored {attempt.score}/{attempt.total_marks} ({attempt.percentage}%)')
            return redirect('exam_results', attempt_pk=attempt.pk)
    
    return redirect('exam_take', exam_pk=attempt.exam.pk)


@login_required
def exam_results(request, attempt_pk):
    """
    Show exam results with explanations (Google Form style) - STUDENT VIEW ONLY
    
    Philosophy: LEARNING-CENTRIC - Show explanations for every question
    Teachers use the separate teacher_view_attempt view
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only. Teachers should use the teacher response dashboard.')
        return redirect('dashboard')
    
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related('exam', 'student'),
        pk=attempt_pk,
        student=request.user
    )
    
    # Get all questions - use the saved randomized order if available
    if attempt.question_order is not None:
        # Retrieve questions in the exact order they were presented during the exam
        questions_dict = {
            q.id: q for q in attempt.exam.questions.prefetch_related('choices').all()
        }
        questions = [questions_dict[qid] for qid in attempt.question_order if qid in questions_dict]
    else:
        # Fallback to original order if question_order wasn't saved (old attempts)
        questions = attempt.exam.questions.prefetch_related('choices').order_by('order')
    
    # Build results data
    results = []
    for question in questions:
        student_answer = StudentAnswer.objects.filter(
            attempt=attempt,
            question=question
        ).select_related('selected_choice').first()
        
        correct_choice = question.choices.filter(is_correct=True).first()
        
        results.append({
            'question': question,
            'student_answer': student_answer,
            'correct_choice': correct_choice,
            'all_choices': question.choices.all()
        })
    
    context = {
        'attempt': attempt,
        'results': results,
        'page_title': f'Results: {attempt.exam.title}'
    }
    return render(request, 'exams/exam_results.html', context)


@login_required
def attempt_history(request):
    """
    Show student's exam attempt history
    
    Philosophy: Track learning progress over time with grouped retakes
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    from collections import defaultdict
    
    # Get all completed attempts
    all_attempts = ExamAttempt.objects.filter(
        student=request.user,
        is_completed=True
    ).select_related('exam', 'student_class').order_by('exam', '-submitted_at')
    
    # Group attempts by exam
    exam_groups = defaultdict(list)
    for attempt in all_attempts:
        exam_groups[attempt.exam.id].append(attempt)
    
    # Build organized data structure
    organized_exams = []
    for exam_id, attempts in exam_groups.items():
        exam = attempts[0].exam
        best_attempt = max(attempts, key=lambda a: a.percentage)
        latest_attempt = attempts[0]  # Already ordered by -submitted_at
        total_attempts = len(attempts)
        passed = any(a.passed for a in attempts)
        
        organized_exams.append({
            'exam': exam,
            'attempts': attempts,
            'best_attempt': best_attempt,
            'latest_attempt': latest_attempt,
            'total_attempts': total_attempts,
            'passed': passed
        })
    
    # Sort by latest attempt date
    organized_exams.sort(key=lambda x: x['latest_attempt'].submitted_at, reverse=True)
    
    # Overall statistics
    total_attempts = all_attempts.count()
    unique_exams = len(organized_exams)
    passed_exams = sum(1 for exam_data in organized_exams if exam_data['passed'])
    avg_score = sum(a.percentage for a in all_attempts) / total_attempts if total_attempts > 0 else 0
    
    context = {
        'organized_exams': organized_exams,
        'total_attempts': total_attempts,
        'unique_exams': unique_exams,
        'passed_exams': passed_exams,
        'avg_score': round(avg_score, 1),
        'page_title': 'My Exam History'
    }
    return render(request, 'exams/attempt_history.html', context)


# ============================================================================
# TEACHER RESPONSE MANAGEMENT VIEWS (Google Forms Style)
# ============================================================================

@login_required
def teacher_exam_responses(request):
    """
    Teacher Response Dashboard: Exam-First Hierarchy
    
    Architecture: Exam → Class/Batch → Student → Attempts → Detailed Responses
    - Each exam appears once, even if used across multiple years
    - Responses are class-scoped (never mix students from different batches)
    - Students grouped under their class
    - Multiple attempts collapsed by default, expanded on interaction
    - Scalable for thousands of responses
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    from django.db.models import Count, Avg, Q, Max, Min
    from collections import defaultdict
    
    # Get teacher's classes
    teacher_classes = StudentClass.objects.filter(
        created_by=request.user,
        is_active=True
    ).prefetch_related('students')
    
    if not teacher_classes.exists():
        context = {
            'no_classes': True,
            'page_title': 'Exam Responses Dashboard'
        }
        return render(request, 'teacher/exam_responses.html', context)
    
    # Filtering parameters
    selected_exam_id = request.GET.get('exam_id')
    selected_class_id = request.GET.get('class_id')
    student_search = request.GET.get('student_search', '').strip()
    pass_filter = request.GET.get('pass_filter')  # 'passed', 'failed', or None
    score_min = request.GET.get('score_min')
    score_max = request.GET.get('score_max')
    sort_by = request.GET.get('sort', 'exam_title')  # exam_title, recent, most_responses
    
    # Build base queryset - ONLY COMPLETED attempts from teacher's classes
    attempts = ExamAttempt.objects.filter(
        student_class__in=teacher_classes,
        is_completed=True
    ).select_related('student', 'exam', 'student_class').order_by('-submitted_at')
    
    # Apply filters
    if selected_exam_id:
        attempts = attempts.filter(exam_id=selected_exam_id)
        selected_exam = Exam.objects.filter(id=selected_exam_id).first()
    else:
        selected_exam = None
    
    if selected_class_id:
        attempts = attempts.filter(student_class_id=selected_class_id)
        selected_class = teacher_classes.filter(id=selected_class_id).first()
    else:
        selected_class = None
    
    if student_search:
        attempts = attempts.filter(
            Q(student__first_name__icontains=student_search) |
            Q(student__last_name__icontains=student_search) |
            Q(student__email__icontains=student_search)
        )
    
    # Get unique exams that have responses
    exams_with_responses = Exam.objects.filter(
        id__in=attempts.values_list('exam_id', flat=True).distinct()
    ).order_by('title')
    
    # Get all exams from teacher's classes for filter dropdown
    all_class_exams = Exam.objects.filter(
        assigned_classes__in=teacher_classes
    ).distinct().order_by('title')
    
    # Build exam-first hierarchy: Exam → Class → Student → Attempts
    exam_hierarchy = []
    
    for exam in exams_with_responses:
        exam_attempts = attempts.filter(exam=exam)
        
        # Get unique classes that have taken this exam
        classes_for_exam = teacher_classes.filter(
            id__in=exam_attempts.values_list('student_class_id', flat=True).distinct()
        )
        
        class_data = []
        for student_class in classes_for_exam:
            class_attempts = exam_attempts.filter(student_class=student_class)
            
            # Group by student within this class
            students_dict = defaultdict(list)
            for attempt in class_attempts:
                students_dict[attempt.student].append(attempt)
            
            student_responses = []
            for student, student_attempts in students_dict.items():
                # Separate test and practice attempts
                test_attempts = [a for a in student_attempts if a.attempt_mode == 'test']
                practice_attempts = [a for a in student_attempts if a.attempt_mode == 'practice']
                
                # Sort attempts by submission time (most recent first)
                test_attempts.sort(key=lambda x: x.submitted_at, reverse=True)
                practice_attempts.sort(key=lambda x: x.submitted_at, reverse=True)
                
                # Combine for display: test attempts first, then practice attempts
                sorted_attempts = test_attempts + practice_attempts
                
                # Find the actual latest attempt by timestamp (for accurate "Latest" badge)
                all_attempts_by_time = sorted(student_attempts, key=lambda x: x.submitted_at, reverse=True)
                actual_latest_attempt = all_attempts_by_time[0] if all_attempts_by_time else None
                
                # Calculate statistics for this student
                valid_attempts = [a for a in sorted_attempts if a.total_marks > 0]
                if not valid_attempts:
                    continue  # Skip students with no valid attempts
                
                best_percentage = max((a.score / a.total_marks * 100) for a in valid_attempts)
                latest_percentage = (actual_latest_attempt.score / actual_latest_attempt.total_marks * 100) if actual_latest_attempt and actual_latest_attempt.total_marks > 0 else 0
                avg_percentage = sum((a.score / a.total_marks * 100) for a in valid_attempts) / len(valid_attempts)
                
                # Apply pass/fail filter
                if pass_filter == 'passed' and not actual_latest_attempt.passed:
                    continue
                if pass_filter == 'failed' and actual_latest_attempt.passed:
                    continue
                
                # Apply score range filter
                if score_min and best_percentage < float(score_min):
                    continue
                if score_max and best_percentage > float(score_max):
                    continue
                
                # Calculate improvement
                improvement = None
                if len(all_attempts_by_time) >= 2:
                    first_percentage = (all_attempts_by_time[-1].score / all_attempts_by_time[-1].total_marks * 100) if all_attempts_by_time[-1].total_marks > 0 else 0
                    improvement = latest_percentage - first_percentage
                
                student_responses.append({
                    'student': student,
                    'attempts': sorted_attempts,
                    'test_attempts': test_attempts,
                    'practice_attempts': practice_attempts,
                    'attempt_count': len(sorted_attempts),
                    'practice_count': len(practice_attempts),
                    'test_count': len(test_attempts),
                    'best_percentage': round(best_percentage, 1),
                    'latest_percentage': round(latest_percentage, 1),
                    'avg_percentage': round(avg_percentage, 1),
                    'improvement': round(improvement, 1) if improvement is not None else None,
                    'latest_attempt': actual_latest_attempt,
                    'latest_attempt_id': actual_latest_attempt.id if actual_latest_attempt else None,
                    'all_passed': all(a.passed for a in sorted_attempts),
                    'any_passed': any(a.passed for a in sorted_attempts),
                })
            
            # Skip class if no students match filters
            if not student_responses:
                continue
            
            # Calculate class statistics
            class_avg = sum(s['avg_percentage'] for s in student_responses) / len(student_responses) if student_responses else 0
            class_pass_rate = (sum(1 for s in student_responses if s['any_passed']) / len(student_responses) * 100) if student_responses else 0
            
            class_data.append({
                'class': student_class,
                'student_responses': student_responses,
                'student_count': len(student_responses),
                'total_attempts': sum(s['attempt_count'] for s in student_responses),
                'class_avg': round(class_avg, 1),
                'class_pass_rate': round(class_pass_rate, 1),
            })
        
        # Skip exam if no classes have matching responses
        if not class_data:
            continue
        
        # Calculate exam-level statistics
        total_students = sum(c['student_count'] for c in class_data)
        total_attempts = sum(c['total_attempts'] for c in class_data)
        exam_avg = sum(c['class_avg'] * c['student_count'] for c in class_data) / total_students if total_students > 0 else 0
        
        exam_hierarchy.append({
            'exam': exam,
            'class_data': class_data,
            'total_students': total_students,
            'total_attempts': total_attempts,
            'exam_avg': round(exam_avg, 1),
            'class_count': len(class_data),
        })
    
    # Apply sorting
    if sort_by == 'exam_title':
        exam_hierarchy.sort(key=lambda x: x['exam'].title)
    elif sort_by == 'recent':
        exam_hierarchy.sort(key=lambda x: x['class_data'][0]['student_responses'][0]['latest_attempt'].submitted_at if x['class_data'] and x['class_data'][0]['student_responses'] else timezone.now(), reverse=True)
    elif sort_by == 'most_responses':
        exam_hierarchy.sort(key=lambda x: x['total_attempts'], reverse=True)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(exam_hierarchy, 10)  # 10 exams per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'teacher_classes': teacher_classes,
        'all_class_exams': all_class_exams,
        'selected_exam': selected_exam,
        'selected_class': selected_class,
        'student_search': student_search,
        'pass_filter': pass_filter,
        'score_min': score_min,
        'score_max': score_max,
        'sort_by': sort_by,
        'exam_count': len(exam_hierarchy),
        'page_title': 'Exam Responses Dashboard'
    }
    return render(request, 'teacher/exam_responses.html', context)


@login_required
def teacher_view_attempt(request, attempt_pk):
    """
    Teacher View of Individual Student Attempt
    
    Philosophy: Detailed response review (like Google Forms individual response)
    - See all student answers with correct/incorrect indicators
    - View explanations and reasoning
    - See time taken, score breakdown
    - Context about the student and class
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related('exam', 'student', 'student_class'),
        pk=attempt_pk
    )
    
    # Verify teacher owns this class
    if not attempt.student_class or attempt.student_class.created_by != request.user:
        messages.error(request, 'Access denied. You can only view responses from your own classes.')
        return redirect('teacher_exam_responses')
    
    # Get all questions - use the saved randomized order if available
    if attempt.question_order is not None:
        # Retrieve questions in the exact order they were presented during the exam
        questions_dict = {
            q.id: q for q in attempt.exam.questions.prefetch_related('choices').all()
        }
        questions = [questions_dict[qid] for qid in attempt.question_order if qid in questions_dict]
    else:
        # Fallback to original order if question_order wasn't saved (old attempts)
        questions = attempt.exam.questions.prefetch_related('choices').order_by('order')
    
    # Build results data
    results = []
    for question in questions:
        student_answer = StudentAnswer.objects.filter(
            attempt=attempt,
            question=question
        ).select_related('selected_choice').first()
        
        correct_choice = question.choices.filter(is_correct=True).first()
        
        results.append({
            'question': question,
            'student_answer': student_answer,
            'correct_choice': correct_choice,
            'all_choices': question.choices.all()
        })
    
    # Get all attempts by this student for this exam (to show retake context)
    all_student_attempts = ExamAttempt.objects.filter(
        student=attempt.student,
        exam=attempt.exam,
        student_class=attempt.student_class,
        is_completed=True
    ).order_by('submitted_at')
    
    attempt_number = list(all_student_attempts).index(attempt) + 1 if attempt in all_student_attempts else 1
    total_attempts_for_exam = all_student_attempts.count()
    
    context = {
        'attempt': attempt,
        'results': results,
        'attempt_number': attempt_number,
        'total_attempts_for_exam': total_attempts_for_exam,
        'is_teacher_view': True,
        'page_title': f"Review: {attempt.student.get_full_name()} - {attempt.exam.title}"
    }
    return render(request, 'teacher/view_attempt.html', context)


# ============================================================================
# TEACHER ANALYTICS VIEWS
# ============================================================================

@login_required
def analytics_dashboard(request):
    """
    Analytics Dashboard: Class-scoped teaching performance metrics
    
    Philosophy: Data-driven insights at class level for better teaching
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    from django.db.models import Avg, Count, Q
    
    # Get teacher's active classes
    teacher_classes = StudentClass.objects.filter(
        created_by=request.user,
        is_active=True
    ).order_by('-created_at')
    
    # Get selected class from query parameter or use first active class
    selected_class_id = request.GET.get('class_id')
    
    if selected_class_id:
        selected_class = get_object_or_404(
            StudentClass,
            pk=selected_class_id,
            created_by=request.user
        )
    elif teacher_classes.exists():
        selected_class = teacher_classes.first()
    else:
        # No classes yet - show empty state
        context = {
            'no_classes': True,
            'page_title': 'Analytics Dashboard'
        }
        return render(request, 'analytics/analytics_dashboard.html', context)
    
    # Get students in selected class only
    class_students = selected_class.students.all()
    
    # Get exams CURRENTLY assigned to this class only
    # CRITICAL: This automatically excludes any exams that were unassigned
    class_exams = selected_class.assigned_exams.all()
    class_exam_ids = list(class_exams.values_list('id', flat=True))
    
    # Get attempts by students in THIS CLASS on exams CURRENTLY assigned to THIS CLASS
    class_attempts = ExamAttempt.objects.filter(
        student__in=class_students,
        exam__in=class_exams,
        exam_id__in=class_exam_ids,  # Explicit filter: only currently assigned exams
        is_completed=True,
        student_class=selected_class
    ).select_related('student', 'exam')
    
    total_attempts = class_attempts.count()
    
    # Calculate class-level statistics with 100% accuracy
    if total_attempts > 0:
        # Calculate average percentage accurately
        total_percentage = 0
        passed_count = 0
        
        for attempt in class_attempts:
            # Get percentage (which is calculated from score/total_marks)
            attempt_percentage = attempt.percentage
            total_percentage += attempt_percentage
            
            # Check if passed (percentage >= exam.pass_percentage)
            if attempt.passed:
                passed_count += 1
        
        # Calculate accurate averages
        avg_percentage = total_percentage / total_attempts
        pass_rate = (passed_count / total_attempts) * 100
        passed_attempts = passed_count
    else:
        avg_percentage = 0
        passed_attempts = 0
        pass_rate = 0
    
    # Recent exams assigned to this class with attempt counts from this class
    recent_exams = []
    for exam in class_exams.order_by('-created_at')[:5]:
        exam_attempts = class_attempts.filter(exam=exam)
        exam.attempt_count = exam_attempts.count()
        recent_exams.append(exam)
    
    # Top performing students in THIS CLASS (using BEST attempt per exam only)
    from collections import defaultdict
    student_performance = defaultdict(lambda: {'exam_percentages': [], 'student': None})
    
    # Get best attempt per student per exam
    for student in class_students:
        for exam in class_exams:
            if exam.total_marks == 0:
                continue
            
            # Get BEST attempt (highest score) for this student on this exam
            # Only from exams currently assigned to this class
            best_attempt = class_attempts.filter(
                student=student,
                exam=exam,
                exam_id__in=class_exam_ids  # Explicit filter: only currently assigned exams
            ).order_by('-score').first()
            
            if best_attempt:
                # Calculate percentage for this exam (cap at 100%)
                exam_percentage = (float(best_attempt.score) / float(exam.total_marks) * 100) if exam.total_marks > 0 else 0
                exam_percentage = min(exam_percentage, 100.0)
                
                student_performance[student.id]['exam_percentages'].append(exam_percentage)
                student_performance[student.id]['student'] = student
    
    top_students = []
    for student_id, data in student_performance.items():
        if data['exam_percentages']:
            # Calculate average percentage across all attempted exams
            avg_percentage = sum(data['exam_percentages']) / len(data['exam_percentages'])
            top_students.append({
                'student': data['student'],
                'avg_score': round(avg_percentage, 1),
                'total_attempts': len(data['exam_percentages'])
            })
    
    top_students.sort(key=lambda x: x['avg_score'], reverse=True)
    top_students = top_students[:5]
    
    context = {
        'teacher_classes': teacher_classes,
        'selected_class': selected_class,
        'total_students': class_students.count(),
        'total_exams': class_exams.count(),
        'assigned_exams_count': class_exams.count(),
        'total_attempts': total_attempts,
        'avg_percentage': round(avg_percentage, 1),
        'pass_rate': round(pass_rate, 1),
        'passed_attempts': passed_attempts,
        'recent_exams': recent_exams,
        'top_students': top_students,
        'page_title': f'Analytics - {selected_class.name}'
    }
    return render(request, 'analytics/analytics_dashboard.html', context)


@login_required
def exam_analytics(request, exam_pk):
    """
    Detailed analytics for a specific exam
    
    Philosophy: Understand how students performed on this assessment
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    from django.db.models import Avg, Count, Q
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    # Get all completed attempts for analytics (no pagination needed for stats)
    all_attempts = ExamAttempt.objects.filter(
        exam=exam,
        is_completed=True
    ).select_related('student')
    
    total_attempts = all_attempts.count()
    
    if total_attempts > 0:
        # Performance statistics (use all_attempts for calculations)
        avg_score = all_attempts.aggregate(Avg('score'))['score__avg'] or 0
        highest_attempt = all_attempts.order_by('-score').first()
        lowest_attempt = all_attempts.order_by('score').first()
        highest_score = highest_attempt.score if highest_attempt else 0
        lowest_score = lowest_attempt.score if lowest_attempt else 0
        
        # Pass/fail analysis (using exam's configured pass percentage)
        pass_threshold = (exam.pass_percentage / 100) * exam.total_marks
        passed = all_attempts.filter(score__gte=pass_threshold).count()
        failed = total_attempts - passed
        pass_rate = (passed / total_attempts * 100) if total_attempts > 0 else 0
        
        # Score distribution (calculate percentage for each attempt)
        score_ranges = {
            '90-100': 0,
            '80-89': 0,
            '70-79': 0,
            '60-69': 0,
            '50-59': 0,
            '40-49': 0,
            '0-39': 0,
        }
        
        for attempt in all_attempts:
            pct = attempt.percentage
            if pct >= 90:
                score_ranges['90-100'] += 1
            elif pct >= 80:
                score_ranges['80-89'] += 1
            elif pct >= 70:
                score_ranges['70-79'] += 1
            elif pct >= 60:
                score_ranges['60-69'] += 1
            elif pct >= 50:
                score_ranges['50-59'] += 1
            elif pct >= 40:
                score_ranges['40-49'] += 1
            else:
                score_ranges['0-39'] += 1
        
        # Average time taken (duration_taken is already in minutes)
        avg_duration = sum([a.duration_taken for a in all_attempts if a.duration_taken]) / total_attempts if total_attempts > 0 else 0
        
    else:
        avg_score = highest_score = lowest_score = pass_rate = avg_duration = 0
        passed = failed = 0
        score_ranges = {}
    
    # Question-wise analysis
    questions = exam.questions.all()
    question_stats = []
    
    for question in questions:
        correct_answers = StudentAnswer.objects.filter(
            question=question,
            is_correct=True
        ).count()
        
        total_answers = StudentAnswer.objects.filter(
            question=question
        ).count()
        
        accuracy = (correct_answers / total_answers * 100) if total_answers > 0 else 0
        
        question_stats.append({
            'question': question,
            'accuracy': round(accuracy, 1),
            'correct': correct_answers,
            'total': total_answers,
            'difficulty': 'Easy' if accuracy >= 70 else 'Medium' if accuracy >= 40 else 'Hard'
        })
    
    # Pagination for Student Attempts table
    # Get attempts ordered by score (highest first) for the table
    attempts_for_table = ExamAttempt.objects.filter(
        exam=exam,
        is_completed=True
    ).select_related('student').order_by('-score', '-submitted_at')
    
    # Pagination setup - 25 attempts per page for optimal performance
    page = request.GET.get('page', 10)
    paginator = Paginator(attempts_for_table, 10)
    
    try:
        attempts_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        attempts_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page
        attempts_page = paginator.page(paginator.num_pages)
    
    context = {
        'exam': exam,
        'attempts': attempts_page,
        'total_attempts': total_attempts,
        'avg_score': round(avg_score, 1),
        'highest_score': round(highest_score, 1),
        'lowest_score': round(lowest_score, 1),
        'pass_rate': round(pass_rate, 1),
        'passed': passed,
        'failed': failed,
        'score_ranges': score_ranges,
        'avg_duration': round(avg_duration, 1),
        'question_stats': question_stats,
        'page_title': f'Analytics - {exam.title}'
    }
    return render(request, 'analytics/exam_analytics.html', context)


@login_required
def student_analytics(request, student_id):
    """
    View individual student's performance across all exams
    
    Philosophy: Track individual learner progress
    Enhanced: Handles thousands of attempts with filtering and pagination
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    from django.db.models import Avg
    from django.core.paginator import Paginator
    
    student = get_object_or_404(User, pk=student_id, role='student')
    
    # SCHOOL-SCOPED: teacher can only view students from their own school
    if student.school != request.user.school:
        messages.error(request, 'Access denied. You can only view students from your school.')
        return redirect('analytics_dashboard')
    
    # Filtering parameters
    selected_class_id = request.GET.get('class_id')
    selected_exam_id = request.GET.get('exam_id')
    date_range = request.GET.get('date_range', 'all')  # all, week, month, year
    sort_by = request.GET.get('sort', 'recent')  # recent, score_high, score_low
    
    # Get all attempts for this student in teacher's classes
    attempts = ExamAttempt.objects.filter(
        student=student,
        student_class__created_by=request.user,
        is_completed=True
    ).select_related('exam', 'student_class')
    
    # Apply filters
    if selected_class_id:
        attempts = attempts.filter(student_class_id=selected_class_id)
        selected_class = StudentClass.objects.filter(id=selected_class_id, created_by=request.user).first()
    else:
        selected_class = None
    
    if selected_exam_id:
        attempts = attempts.filter(exam_id=selected_exam_id)
        selected_exam = Exam.objects.filter(id=selected_exam_id).first()
    else:
        selected_exam = None
    
    # Date range filtering
    if date_range == 'week':
        from datetime import timedelta
        week_ago = timezone.now() - timedelta(days=7)
        attempts = attempts.filter(submitted_at__gte=week_ago)
    elif date_range == 'month':
        from datetime import timedelta
        month_ago = timezone.now() - timedelta(days=30)
        attempts = attempts.filter(submitted_at__gte=month_ago)
    elif date_range == 'year':
        from datetime import timedelta
        year_ago = timezone.now() - timedelta(days=365)
        attempts = attempts.filter(submitted_at__gte=year_ago)
    
    # Apply sorting
    if sort_by == 'recent':
        attempts = attempts.order_by('-submitted_at')
    elif sort_by == 'score_high':
        attempts = attempts.order_by('-score', '-submitted_at')
    elif sort_by == 'score_low':
        attempts = attempts.order_by('score', '-submitted_at')
    else:
        attempts = attempts.order_by('-submitted_at')
    
    total_attempts = attempts.count()
    
    # Statistics
    if total_attempts > 0:
        # Calculate manually since percentage is a property
        total_percentage = sum((a.score / a.total_marks * 100) for a in attempts if a.total_marks > 0)
        avg_percentage = total_percentage / total_attempts if total_attempts > 0 else 0
        
        highest_score = max((a.percentage for a in attempts), default=0)
        lowest_score = min((a.percentage for a in attempts), default=0)
        passed = sum(1 for a in attempts if a.passed)
        pass_rate = (passed / total_attempts * 100) if total_attempts > 0 else 0
    else:
        avg_percentage = highest_score = lowest_score = pass_rate = 0
        passed = 0
    
    # Subject-wise performance (grouped summary)
    subject_performance = {}
    for attempt in attempts:
        subject = attempt.exam.subject
        if subject not in subject_performance:
            subject_performance[subject] = {
                'total': 0,
                'sum': 0,
                'passed': 0
            }
        subject_performance[subject]['total'] += 1
        subject_performance[subject]['sum'] += attempt.percentage
        if attempt.passed:
            subject_performance[subject]['passed'] += 1
    
    # Calculate averages
    for subject, data in subject_performance.items():
        data['avg'] = round(data['sum'] / data['total'], 1)
        data['pass_rate'] = round((data['passed'] / data['total'] * 100), 1) if data['total'] > 0 else 0
    
    # Get teacher's classes for filter dropdown
    teacher_classes = StudentClass.objects.filter(
        created_by=request.user,
        students=student,
        is_active=True
    )
    
    # Get exams for filter dropdown
    class_exams = Exam.objects.filter(
        assigned_classes__in=teacher_classes
    ).distinct().order_by('title')
    
    # Pagination
    paginator = Paginator(attempts, 20)  # 20 attempts per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'student': student,
        'page_obj': page_obj,
        'total_attempts': total_attempts,
        'avg_percentage': round(avg_percentage, 1),
        'highest_score': round(highest_score, 1),
        'lowest_score': round(lowest_score, 1),
        'pass_rate': round(pass_rate, 1),
        'passed': passed,
        'subject_performance': subject_performance,
        'teacher_classes': teacher_classes,
        'class_exams': class_exams,
        'selected_class': selected_class,
        'selected_exam': selected_exam,
        'date_range': date_range,
        'sort_by': sort_by,
        'page_title': f'Analytics - {student.get_full_name()}'
    }
    return render(request, 'analytics/student_analytics.html', context)


@login_required
def question_analytics(request, exam_pk):
    """
    Question difficulty analysis for an exam
    
    Philosophy: Identify knowledge gaps and improve questions
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
    
    questions = exam.questions.all()
    question_details = []
    
    for question in questions:
        # Get all answers for this question
        all_answers = StudentAnswer.objects.filter(question=question)
        total_answers = all_answers.count()
        
        if total_answers > 0:
            correct_answers = all_answers.filter(is_correct=True).count()
            accuracy = (correct_answers / total_answers * 100)
            
            # Choice distribution
            choice_stats = []
            for choice in question.choices.all():
                selected_count = all_answers.filter(selected_choice=choice).count()
                percentage = (selected_count / total_answers * 100) if total_answers > 0 else 0
                choice_stats.append({
                    'choice': choice,
                    'selected_count': selected_count,
                    'percentage': round(percentage, 1)
                })
            
            question_details.append({
                'question': question,
                'total_responses': total_answers,
                'correct_responses': correct_answers,
                'accuracy': round(accuracy, 1),
                'difficulty': 'Easy' if accuracy >= 70 else 'Medium' if accuracy >= 40 else 'Hard',
                'choice_stats': choice_stats
            })
        else:
            question_details.append({
                'question': question,
                'total_responses': 0,
                'correct_responses': 0,
                'accuracy': 0,
                'difficulty': 'Not attempted',
                'choice_stats': []
            })
    
    # Calculate difficulty summary counts
    easy_count = sum(1 for q in question_details if q['difficulty'] == 'Easy')
    medium_count = sum(1 for q in question_details if q['difficulty'] == 'Medium')
    hard_count = sum(1 for q in question_details if q['difficulty'] == 'Hard')
    
    context = {
        'exam': exam,
        'question_details': question_details,
        'easy_count': easy_count,
        'medium_count': medium_count,
        'hard_count': hard_count,
        'page_title': f'Question Analysis - {exam.title}'
    }
    return render(request, 'analytics/question_analytics.html', context)


# ============================================================================
# CLASS MANAGEMENT VIEWS
# ============================================================================

@login_required
def class_list(request):
    """
    List all classes created by teacher
    
    Philosophy: Organize students into manageable cohorts
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    classes = StudentClass.objects.filter(
        created_by=request.user
    ).prefetch_related('students', 'assigned_exams')
    
    # Calculate statistics
    total_classes = classes.count()
    active_classes = classes.filter(is_active=True).count()
    total_students = sum(c.student_count for c in classes)
    
    context = {
        'classes': classes,
        'total_classes': total_classes,
        'active_classes': active_classes,
        'total_students': total_students,
        'page_title': 'My Classes'
    }
    return render(request, 'classes/class_list.html', context)


@login_required
def class_create(request):
    """Create a new class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        year = request.POST.get('year')
        
        if not name or not year:
            messages.error(request, 'Class name and year are required.')
            return redirect('class_create')
        
        try:
            student_class = StudentClass.objects.create(
                name=name,
                description=description,
                year=int(year),
                created_by=request.user,
                school=request.user.school   # SCHOOL-SCOPED
            )
            messages.success(request, f'Class "{student_class.name}" created successfully!')
            return redirect('class_detail', class_pk=student_class.pk)
        except Exception as e:
            messages.error(request, f'Error creating class: {str(e)}')
            return redirect('class_create')
    
    from datetime import datetime
    current_year = datetime.now().year
    
    context = {
        'current_year': current_year,
        'page_title': 'Create Class'
    }
    return render(request, 'classes/class_form.html', context)


@login_required
def class_detail(request, class_pk):
    """View class details with students and exams"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    # Get students in this class
    students = student_class.students.all()
    
    # Get assigned exams
    assigned_exams = student_class.assigned_exams.all()
    
    # Get available students (not in this class yet) — SCHOOL-SCOPED
    available_students = User.objects.filter(
        role='student',
        school=student_class.school   # Only students from same school
    ).exclude(
        id__in=students.values_list('id', flat=True)
    )

    # Get available exams (created by this teacher, not assigned yet) — SCHOOL-SCOPED
    available_exams = Exam.objects.filter(
        created_by=request.user,
        school=student_class.school   # Only exams from same school
    ).exclude(
        id__in=assigned_exams.values_list('id', flat=True)
    )
    
    context = {
        'student_class': student_class,
        'students': students,
        'assigned_exams': assigned_exams,
        # We no longer pass all available_students to avoid large payloads.
        # They will be fetched via AJAX search instead.
        'available_exams': available_exams,
        'page_title': f'{student_class.name}'
    }
    return render(request, 'classes/class_detail.html', context)


from django.http import JsonResponse

@login_required
def search_students_for_class(request, class_pk):
    """
    AJAX endpoint to search for students within the teacher's school 
    who are not already in the specified class.
    """
    if not request.user.is_teacher:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    query = request.GET.get('q', '').strip()
    if not query or len(query) < 2:
        return JsonResponse({'results': []})
        
    # Strictly scoped to school, and exclude already enrolled students
    enrolled_ids = student_class.students.values_list('id', flat=True)
    
    # Filter by name or username
    available = User.objects.filter(
        role='student',
        school=student_class.school
    ).exclude(
        id__in=enrolled_ids
    ).filter(
        Q(first_name__icontains=query) | 
        Q(last_name__icontains=query) | 
        Q(username__icontains=query)
    ).order_by('first_name', 'last_name')[:20]  # Limit to 20 results for speed
    
    results = []
    for student in available:
        results.append({
            'id': student.id,
            'name': student.get_full_name() or student.username,
            'username': student.username,
            'initials': (student.first_name[:1] if student.first_name else student.username[:1]).upper()
        })
        
    return JsonResponse({'results': results})


@login_required
def class_update(request, class_pk):
    """Update class details"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        student_class.name = request.POST.get('name', student_class.name)
        student_class.description = request.POST.get('description', '')
        student_class.year = int(request.POST.get('year', student_class.year))
        student_class.is_active = request.POST.get('is_active') == 'on'
        student_class.save()
        
        messages.success(request, f'Class "{student_class.name}" updated successfully!')
        return redirect('class_detail', class_pk=student_class.pk)
    
    context = {
        'student_class': student_class,
        'page_title': f'Edit {student_class.name}'
    }
    return render(request, 'classes/class_form.html', context)


@login_required
def class_delete(request, class_pk):
    """Delete a class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        class_name = student_class.name
        student_class.delete()
        messages.success(request, f'Class "{class_name}" deleted successfully!')
        return redirect('class_list')
    
    context = {
        'student_class': student_class,
        'page_title': f'Delete {student_class.name}'
    }
    return render(request, 'classes/class_confirm_delete.html', context)


@login_required
def class_add_students(request, class_pk):
    """Add students to a class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        student_ids = request.POST.getlist('students')
        if student_ids:
            # SCHOOL-SCOPED: only students from the same school can be added
            students = User.objects.filter(
                id__in=student_ids,
                role='student',
                school=student_class.school
            )
            added_count = 0
            already_enrolled_students = []

            for student in students:
                existing_class = student.student_classes.filter(is_active=True).first()
                if existing_class:
                    already_enrolled_students.append(
                        f'{student.get_full_name() or student.username} (enrolled in {existing_class.name})'
                    )
                else:
                    student_class.students.add(student)
                    added_count += 1

            if added_count > 0:
                messages.success(request, f'{added_count} student(s) added to {student_class.name}')
            if already_enrolled_students:
                for student_info in already_enrolled_students:
                    messages.warning(request, f'{student_info} - Student is already enrolled in a class and cannot join multiple classes.')
        else:
            messages.warning(request, 'No students selected.')
        return redirect('class_detail', class_pk=class_pk)
    return redirect('class_detail', class_pk=class_pk)


@login_required
def class_remove_student(request, class_pk, student_id):
    """Remove a student from a class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        student = get_object_or_404(User, id=student_id, role='student')
        student_class.students.remove(student)
        messages.success(request, f'{student.get_full_name() or student.email} removed from {student_class.name}')
    
    return redirect('class_detail', class_pk=class_pk)


@login_required
def class_assign_exam(request, class_pk):
    """Assign exam to a class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        exam_ids = request.POST.getlist('exams')
        if exam_ids:
            # SCHOOL-SCOPED: only exams from same school can be assigned
            exams = Exam.objects.filter(
                id__in=exam_ids,
                created_by=request.user,
                school=student_class.school
            )
            for exam in exams:
                exam.assigned_classes.add(student_class)
            messages.success(request, f'{exams.count()} exam(s) assigned to {student_class.name}')
        else:
            messages.warning(request, 'No exams selected.')
        return redirect('class_detail', class_pk=class_pk)
    return redirect('class_detail', class_pk=class_pk)


@login_required
def class_unassign_exam(request, class_pk, exam_pk):
    """Unassign exam from a class"""
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass, 
        pk=class_pk, 
        created_by=request.user
    )
    
    if request.method == 'POST':
        exam = get_object_or_404(Exam, pk=exam_pk, created_by=request.user)
        exam.assigned_classes.remove(student_class)
        messages.success(request, f'{exam.title} unassigned from {student_class.name}')
    
    return redirect('class_detail', class_pk=class_pk)


# ============================================================================
# ENROLLMENT VIEWS
# ============================================================================

@login_required
def student_enrollment(request):
    """
    Student enrollment dashboard - browse and join classes
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Get student's current classes
    enrolled_classes = request.user.student_classes.filter(is_active=True)
    
    # Get available classes for enrollment (school-scoped + self-enrollment enabled)
    available_classes = StudentClass.objects.filter(
        is_active=True,
        allow_self_enrollment=True,
        school=request.user.school   # SCHOOL-SCOPED: only same-school classes
    ).exclude(
        students=request.user
    ).select_related('created_by')
    
    context = {
        'enrolled_classes': enrolled_classes,
        'available_classes': available_classes,
        'page_title': 'My Classes'
    }
    return render(request, 'enrollment/student_enrollment.html', context)


@login_required
def join_class_by_code(request):
    """
    Join a class using enrollment code
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        
        if not code:
            messages.error(request, 'Please enter a class code.')
            return redirect('student_enrollment')
        
        try:
            student_class = StudentClass.objects.get(
                enrollment_code=code,
                is_active=True,
                allow_self_enrollment=True
            )
            
            # Check if already enrolled in this class
            if request.user in student_class.students.all():
                messages.info(request, f'You are already enrolled in {student_class.name}.')
                return redirect('student_enrollment')
            
            # Check if student is already enrolled in any other class
            existing_class = request.user.student_classes.filter(is_active=True).first()
            
            if existing_class:
                messages.warning(
                    request, 
                    f'You cannot join {student_class.name} because you are already enrolled in {existing_class.name}. '
                    f'A student cannot join more than one class.'
                )
            else:
                # Student can join the class
                student_class.students.add(request.user)
                messages.success(request, f'Successfully joined {student_class.name}!')
            
            return redirect('student_enrollment')
            
        except StudentClass.DoesNotExist:
            messages.error(request, 'Invalid class code or class not available for enrollment.')
            return redirect('student_enrollment')
    
    return redirect('student_enrollment')


@login_required
def leave_class(request, class_pk):
    """
    Leave a class
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(StudentClass, pk=class_pk)
    
    if request.user in student_class.students.all():
        student_class.students.remove(request.user)
        messages.success(request, f'You have left {student_class.name}.')
    else:
        messages.info(request, 'You are not enrolled in this class.')
    
    return redirect('student_enrollment')


@login_required
def toggle_self_enrollment(request, class_pk):
    """
    Teacher: Toggle self-enrollment for a class
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass,
        pk=class_pk,
        created_by=request.user
    )
    
    if student_class.allow_self_enrollment:
        student_class.disable_self_enrollment()
        messages.info(request, f'Self-enrollment disabled for {student_class.name}.')
    else:
        student_class.enable_self_enrollment()
        messages.success(request, f'Self-enrollment enabled for {student_class.name}. Code: {student_class.enrollment_code}')
    
    return redirect('class_detail', class_pk=class_pk)


@login_required
def regenerate_enrollment_code(request, class_pk):
    """
    Teacher: Regenerate enrollment code for a class
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass,
        pk=class_pk,
        created_by=request.user
    )
    
    old_code = student_class.enrollment_code
    new_code = student_class.generate_enrollment_code()
    
    messages.success(request, f'New enrollment code generated: {new_code}')
    
    return redirect('class_detail', class_pk=class_pk)


@login_required
def class_enrollment_settings(request, class_pk):
    """
    Teacher: Manage enrollment settings for a class
    """
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Teachers only.')
        return redirect('dashboard')
    
    student_class = get_object_or_404(
        StudentClass,
        pk=class_pk,
        created_by=request.user
    )
    
    context = {
        'student_class': student_class,
        'page_title': f'Enrollment Settings - {student_class.name}'
    }
    return render(request, 'enrollment/enrollment_settings.html', context)


# =====================================
# STUDENT PERFORMANCE PROGRESS TRACKING
# =====================================
# Philosophy: Progress is ALWAYS scoped to something specific.
# NEVER mix different exams, subjects, or domains.
# Answer one question: "Is the student improving in THIS specific thing over time?"

@login_required
def student_progress_dashboard(request):
    """
    Student Progress Dashboard - Entry Point
    
    Shows two scoped categories (NO global metrics):
    1. Track by Exam (exam-wise progress)
    2. Track by Subject (subject-wise progress)
    
    This is NOT a summary view - it's a navigation hub to scoped progress.
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Get all completed attempts for this student
    completed_attempts = ExamAttempt.objects.filter(
        student=request.user,
        is_completed=True
    ).select_related('exam', 'student_class').order_by('-submitted_at')
    
    # Group exams by: 1) Exam (for retakes), 2) Subject
    # Build structure for the two tracking categories
    
    # 1. EXAM-WISE: Group by specific exam (shows retakes)
    exam_data = {}
    for attempt in completed_attempts:
        exam_id = attempt.exam.id
        if exam_id not in exam_data:
            exam_data[exam_id] = {
                'exam': attempt.exam,
                'attempt_count': 0,
                'latest_attempt': None,
                'first_attempt': None,
                'best_score': 0,
                'latest_score': 0,
            }
        
        exam_data[exam_id]['attempt_count'] += 1
        exam_data[exam_id]['best_score'] = max(exam_data[exam_id]['best_score'], attempt.percentage)
        
        if not exam_data[exam_id]['latest_attempt'] or attempt.submitted_at > exam_data[exam_id]['latest_attempt'].submitted_at:
            exam_data[exam_id]['latest_attempt'] = attempt
            exam_data[exam_id]['latest_score'] = attempt.percentage
        
        if not exam_data[exam_id]['first_attempt'] or attempt.submitted_at < exam_data[exam_id]['first_attempt'].submitted_at:
            exam_data[exam_id]['first_attempt'] = attempt
    
    # Calculate improvement for each exam
    for exam_id, data in exam_data.items():
        if data['attempt_count'] > 1:
            first_score = data['first_attempt'].percentage
            latest_score = data['latest_score']
            data['improvement'] = round(latest_score - first_score, 2)
        else:
            data['improvement'] = 0
    
    # 2. SUBJECT-WISE: Group by subject
    subject_data = {}
    for attempt in completed_attempts:
        subject = attempt.exam.subject
        if subject not in subject_data:
            subject_data[subject] = {
                'subject': subject,
                'exam_count': set(),
                'total_attempts': 0,
                'latest_attempt': None,
            }
        
        subject_data[subject]['exam_count'].add(attempt.exam.id)
        subject_data[subject]['total_attempts'] += 1
        
        if not subject_data[subject]['latest_attempt'] or attempt.submitted_at > subject_data[subject]['latest_attempt'].submitted_at:
            subject_data[subject]['latest_attempt'] = attempt
    
    # Convert sets to counts
    for subject, data in subject_data.items():
        data['exam_count'] = len(data['exam_count'])
    
    context = {
        'page_title': 'My Progress Tracking',
        'exam_data': sorted(exam_data.values(), key=lambda x: x['latest_attempt'].submitted_at, reverse=True),
        'subject_data': sorted(subject_data.values(), key=lambda x: x['subject']),
        'total_completed_attempts': completed_attempts.count(),
    }
    
    return render(request, 'progress/progress_dashboard.html', context)


@login_required
def exam_progress(request, exam_id):
    """
    Exam-Specific Progress Tracking
    
    Shows ONLY attempts for THIS specific exam.
    Answers: "Am I improving in THIS exam over time?"
    
    Features:
    - All attempts/retakes for this exam
    - Score progression chart
    - Attempt comparison table
    - Improvement indicators
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    exam = get_object_or_404(Exam, pk=exam_id)
    
    # Get ALL attempts for THIS exam only (never mix with other exams)
    attempts = ExamAttempt.objects.filter(
        student=request.user,
        exam=exam,
        is_completed=True
    ).select_related('student_class').order_by('submitted_at')
    
    if not attempts.exists():
        messages.info(request, 'You have not completed this exam yet.')
        return redirect('student_progress_dashboard')
    
    # Calculate statistics for THIS exam only
    attempt_list = list(attempts)
    first_attempt = attempt_list[0]
    latest_attempt = attempt_list[-1]
    best_attempt = max(attempt_list, key=lambda a: a.percentage)
    
    scores = [float(attempt.percentage) for attempt in attempt_list]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    # Calculate improvement
    if len(attempt_list) > 1:
        improvement = latest_attempt.percentage - first_attempt.percentage
        improvement_percent = (improvement / first_attempt.percentage * 100) if first_attempt.percentage > 0 else 0
    else:
        improvement = 0
        improvement_percent = 0
    
    # Attempt comparison data
    comparison_data = []
    for i, attempt in enumerate(attempt_list):
        comparison_data.append({
            'attempt_number': i + 1,
            'attempt': attempt,
            'improvement_from_previous': attempt.percentage - attempt_list[i-1].percentage if i > 0 else 0,
            'improvement_from_first': attempt.percentage - first_attempt.percentage,
        })
    
    context = {
        'page_title': f'Progress: {exam.title}',
        'exam': exam,
        'attempts': attempt_list,
        'attempt_count': len(attempt_list),
        'first_attempt': first_attempt,
        'latest_attempt': latest_attempt,
        'best_attempt': best_attempt,
        'avg_score': round(avg_score, 2),
        'improvement': round(improvement, 2),
        'improvement_percent': round(improvement_percent, 2),
        'comparison_data': comparison_data,
        'is_improving': improvement > 0,
    }
    
    return render(request, 'progress/exam_progress.html', context)


@login_required
def subject_progress(request, subject):
    """
    Subject-Specific Progress Tracking
    
    Shows ONLY exams within THIS subject.
    Answers: "Am I improving in THIS subject over time?"
    
    NEVER mixes different subjects.
    Shows chronological performance across all exams in this subject.
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. Students only.')
        return redirect('dashboard')
    
    # Get ALL attempts for exams in THIS subject only
    attempts = ExamAttempt.objects.filter(
        student=request.user,
        exam__subject=subject,
        is_completed=True
    ).select_related('exam', 'student_class').order_by('submitted_at')
    
    if not attempts.exists():
        messages.info(request, f'You have not completed any exams in {subject} yet.')
        return redirect('student_progress_dashboard')
    
    # Group attempts by exam within this subject
    exam_data = {}
    for attempt in attempts:
        exam_id = attempt.exam.id
        if exam_id not in exam_data:
            exam_data[exam_id] = {
                'exam': attempt.exam,
                'attempts': [],
                'best_score': 0,
                'latest_score': 0,
                'attempt_count': 0,
                'latest_attempt': None,
            }
        
        exam_data[exam_id]['attempts'].append(attempt)
        exam_data[exam_id]['best_score'] = max(exam_data[exam_id]['best_score'], attempt.percentage)
        exam_data[exam_id]['latest_score'] = attempt.percentage
        exam_data[exam_id]['attempt_count'] += 1
        exam_data[exam_id]['latest_attempt'] = attempt  # Always update to the latest one
    
    context = {
        'page_title': f'Progress: {subject}',
        'subject': subject,
        'exam_data': sorted(exam_data.values(), key=lambda x: x['attempts'][-1].submitted_at, reverse=True),
        'total_attempts': attempts.count(),
        'unique_exams': len(exam_data),
        'attempts': list(attempts),
    }
    
    return render(request, 'progress/subject_progress.html', context)


@login_required
def class_leaderboard(request, class_id):
    """
    Class-Scoped Leaderboard
    
    Ranks students based on cumulative performance across ALL exams assigned to this class.
    
    Rules:
    - Only the BEST attempt per exam counts
    - Only completed attempts are considered
    - Each student appears once
    - Score = Sum of best attempts from all assigned exams
    
    Strict class-scoping: NEVER mixes students from different classes.
    """
    student_class = get_object_or_404(StudentClass, pk=class_id)
    
    # SCHOOL-SCOPED: block cross-school leaderboard access
    if request.user.is_authenticated and not request.user.is_superuser:
        if student_class.school != request.user.school:
            messages.error(request, 'Access denied. You cannot view leaderboards from another school.')
            return redirect('dashboard')
    
    # Verify access: Teachers can view any class in their school, students only their own class
    if request.user.is_student:
        if not student_class.students.filter(id=request.user.id).exists():
            messages.error(request, 'You can only view the leaderboard for your own class.')
            return redirect('dashboard')
    elif not request.user.is_teacher:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Get all teacher's classes for the dropdown (only for teachers)
    teacher_classes = None
    if request.user.is_teacher:
        teacher_classes = StudentClass.objects.filter(
            created_by=request.user,
            is_active=True
        ).order_by('name')
    
    # Get all ACTIVE exams CURRENTLY assigned to THIS class only
    # CRITICAL: This automatically excludes any exams that were unassigned from the class
    # Both 'test' and 'practice_test' types are included
    # Only Test attempts count towards leaderboard (Practice attempts excluded)
    assigned_exams = student_class.assigned_exams.all()
    
    # Defensive check: Get list of currently assigned exam IDs for explicit filtering
    assigned_exam_ids = list(assigned_exams.values_list('id', flat=True))
    
    if not assigned_exams.exists():
        context = {
            'page_title': f'Leaderboard: {student_class.name}',
            'student_class': student_class,
            'leaderboard_data': [],
            'total_exams': 0,
            'teacher_classes': teacher_classes,
        }
        return render(request, 'leaderboard/class_leaderboard.html', context)
    
    # Get all ACTIVE students enrolled in this class
    students = student_class.students.filter(role='student', is_active=True)
    
    if not students.exists():
        context = {
            'page_title': f'Leaderboard: {student_class.name}',
            'student_class': student_class,
            'leaderboard_data': [],
            'total_exams': assigned_exams.count(),
            'teacher_classes': teacher_classes,
        }
        return render(request, 'leaderboard/class_leaderboard.html', context)
    
    # Calculate cumulative score for each student
    leaderboard_data = []
    
    for student in students:
        total_percentage_sum = 0  # Sum of percentages across all attempted exams
        exams_attempted = 0
        exam_details = []
        
        # For each assigned exam, get the student's BEST attempt
        for exam in assigned_exams:
            # Edge case: Ensure exam has questions and total_marks > 0
            if exam.total_marks == 0:
                continue
            
            # Get BEST TEST attempt (highest score) for this specific exam in this specific class
            # CRITICAL: Double-check exam is currently assigned (defensive programming)
            # Only 'test' mode attempts count for leaderboard
            best_attempt = ExamAttempt.objects.filter(
                student=student,
                exam=exam,
                exam_id__in=assigned_exam_ids,  # Explicit filter: only currently assigned exams
                student_class=student_class,
                is_completed=True,
                attempt_mode='test'  # Only test attempts count for leaderboard
            ).order_by('-score').first()
            
            if best_attempt:
                # Calculate percentage for this exam (cap at 100% to handle data anomalies)
                exam_percentage = (float(best_attempt.score) / float(exam.total_marks) * 100) if exam.total_marks > 0 else 0
                exam_percentage = min(exam_percentage, 100.0)  # Cap at 100%
                
                # Cap displayed score at total marks to handle data anomalies
                displayed_score = min(float(best_attempt.score), float(exam.total_marks))
                
                # Add percentage to cumulative sum
                total_percentage_sum += exam_percentage
                exams_attempted += 1
                
                # Calculate time in seconds for precision
                if best_attempt.submitted_at:
                    time_taken_seconds = (best_attempt.submitted_at - best_attempt.started_at).total_seconds()
                else:
                    time_taken_seconds = 0
                
                # Format individual exam time as mm:ss with proper rounding
                # Round to nearest second for display accuracy
                rounded_seconds = round(time_taken_seconds)
                time_minutes = int(rounded_seconds // 60)
                time_seconds = int(rounded_seconds % 60)
                time_taken_formatted = f"{time_minutes}:{time_seconds:02d}"
                
                # Store exam details for breakdown
                exam_details.append({
                    'exam': exam,
                    'score': displayed_score,
                    'total': exam.total_marks,
                    'percentage': round(exam_percentage, 2),
                    'time_taken_seconds': time_taken_seconds,  # Time in seconds for precision
                    'time_taken': time_taken_formatted,  # Formatted time (mm:ss) for display
                })
        
        # Only include students who have attempted at least one exam
        if exams_attempted > 0:
            # Calculate overall average percentage across all attempted exams
            overall_percentage = total_percentage_sum / exams_attempted
            
            # LEADERBOARD FILTER: Only include students with average score >= 75%
            if overall_percentage >= 75.0:
                # Calculate average completion time in seconds across all test attempts
                total_time_seconds = sum(detail['time_taken_seconds'] for detail in exam_details)
                average_time_seconds = total_time_seconds / exams_attempted if exams_attempted > 0 else 0
                
                # Format as mm:ss with proper rounding for accuracy
                # Round to nearest second before formatting
                rounded_avg_seconds = round(average_time_seconds)
                avg_minutes = int(rounded_avg_seconds // 60)
                avg_seconds = int(rounded_avg_seconds % 60)
                average_time_formatted = f"{avg_minutes}:{avg_seconds:02d}"
                
                leaderboard_data.append({
                    'student': student,
                    'average_percentage': round(overall_percentage, 2),
                    'average_time_seconds': average_time_seconds,  # For sorting
                    'average_time': average_time_formatted,  # For display (mm:ss)
                    'exams_attempted': exams_attempted,
                    'exam_details': exam_details,
                })
    
    # Sort by ranking priority (all descending for negated values, ascending for time):
    # 1. Number of Test exams taken (more tests = higher rank)
    # 2. Average Test score percentage (higher percentage = higher rank)
    # 3. Average Test completion time (lower time = higher rank)
    leaderboard_data.sort(key=lambda x: (-x['exams_attempted'], -x['average_percentage'], x['average_time_seconds']))
    
    # Add rank with proper tie handling (considering all three criteria)
    # Students with same exams_attempted, average_percentage, AND average_time get the same rank
    for i, data in enumerate(leaderboard_data):
        if i == 0:
            data['rank'] = 1
        else:
            prev_data = leaderboard_data[i - 1]
            if (data['exams_attempted'] == prev_data['exams_attempted'] and
                data['average_percentage'] == prev_data['average_percentage'] and 
                data['average_time_seconds'] == prev_data['average_time_seconds']):
                # Same on all three criteria = same rank
                data['rank'] = prev_data['rank']
            else:
                # Different on any criteria = next sequential rank
                data['rank'] = i + 1    
    context = {
        'page_title': f'Leaderboard: {student_class.name}',
        'student_class': student_class,
        'leaderboard_data': leaderboard_data,
        'total_exams': assigned_exams.count(),
        'teacher_classes': teacher_classes,
    }
    
    return render(request, 'leaderboard/class_leaderboard.html', context)


@login_required
def student_rankings(request):
    """
    Display a dedicated leaderboard page for students with class selection.
    Students can toggle between their enrolled classes to view different leaderboards.
    Teachers are redirected to their latest created class leaderboard.
    """
    # If teacher, redirect to the latest created class leaderboard
    if request.user.is_teacher:
        latest_class = StudentClass.objects.filter(
            created_by=request.user,
            is_active=True
        ).order_by('-created_at').first()
        
        if latest_class:
            return redirect('class_leaderboard', class_id=latest_class.id)
        else:
            messages.info(request, 'No classes found. Create a class first to view leaderboards.')
            return redirect('class_list')
    
    # Get all classes the student is enrolled in
    enrolled_classes = StudentClass.objects.filter(
        students=request.user,
        students__role='student',
        students__is_active=True
    ).order_by('name')
    
    # If no enrolled classes, show message
    if not enrolled_classes.exists():
        return render(request, 'leaderboard/student_rankings.html', {
            'page_title': 'Class Rankings',
            'enrolled_classes': [],
            'selected_class': None,
            'leaderboard_data': [],
        })
    
    # Get selected class (from query param or default to first)
    selected_class_id = request.GET.get('class_id')
    if selected_class_id:
        try:
            selected_class = enrolled_classes.get(id=selected_class_id)
        except StudentClass.DoesNotExist:
            # Fallback to first class if invalid ID
            selected_class = enrolled_classes.first()
    else:
        # Default to first enrolled class
        selected_class = enrolled_classes.first()
    
    # Get leaderboard data for selected class (reuse existing logic)
    # Both 'test' and 'practice_test' exam types are included
    # Only Test mode attempts count towards leaderboard (Practice mode attempts excluded)
    assigned_exams = selected_class.assigned_exams.all()
    assigned_exam_ids = list(assigned_exams.values_list('id', flat=True))
    
    if not assigned_exams.exists():
        return render(request, 'leaderboard/student_rankings.html', {
            'page_title': 'Class Rankings',
            'enrolled_classes': enrolled_classes,
            'selected_class': selected_class,
            'leaderboard_data': [],
            'total_exams': 0,
            'total_students': 0,
            'participating_students': 0,
        })
    
    # Get all active students in the class
    students = selected_class.students.filter(role='student', is_active=True)
    
    # Calculate scores for each student
    leaderboard_data = []
    for student in students:
        total_percentage_sum = 0
        exams_attempted = 0
        total_time_seconds = 0
        exam_details = []
        
        for exam in assigned_exams:
            # Skip exams with zero marks to avoid division errors
            if exam.total_marks == 0:
                continue
                
            # Get student's BEST TEST attempt for this exam
            # CRITICAL: Only from exams currently assigned to this class
            # Only 'test' mode attempts count for leaderboard
            best_attempt = ExamAttempt.objects.filter(
                student=student,
                exam=exam,
                exam_id__in=assigned_exam_ids,  # Explicit filter: only currently assigned exams
                student_class=selected_class,
                is_completed=True,
                attempt_mode='test'  # Only test attempts count for leaderboard
            ).order_by('-score').first()
            
            if best_attempt:
                # Calculate percentage for this exam (cap at 100% to handle data anomalies)
                exam_percentage = (float(best_attempt.score) / float(exam.total_marks) * 100) if exam.total_marks > 0 else 0
                exam_percentage = min(exam_percentage, 100.0)  # Cap at 100%
                
                # Cap displayed score at total marks to handle data anomalies
                displayed_score = min(float(best_attempt.score), float(exam.total_marks))
                
                # Calculate time in seconds for precision
                if best_attempt.submitted_at:
                    time_taken_seconds = (best_attempt.submitted_at - best_attempt.started_at).total_seconds()
                else:
                    time_taken_seconds = 0
                
                # Format individual exam time as mm:ss with proper rounding
                rounded_seconds = round(time_taken_seconds)
                time_minutes = int(rounded_seconds // 60)
                time_seconds = int(rounded_seconds % 60)
                time_taken_formatted = f"{time_minutes}:{time_seconds:02d}"
                
                # Add percentage to cumulative sum
                total_percentage_sum += exam_percentage
                total_time_seconds += time_taken_seconds
                exams_attempted += 1
                
                exam_details.append({
                    'exam': exam,
                    'score': displayed_score,
                    'total': exam.total_marks,
                    'percentage': round(exam_percentage, 2),
                    'time_taken_seconds': time_taken_seconds,
                    'time_taken': time_taken_formatted,
                })
        
        # Only include students who have completed at least one exam
        if exams_attempted > 0:
            average_percentage = total_percentage_sum / exams_attempted
            
            # LEADERBOARD FILTER: Only include students with average score >= 75%
            if average_percentage >= 75.0:
                # Calculate average completion time in seconds across all test attempts
                average_time_seconds = total_time_seconds / exams_attempted if exams_attempted > 0 else 0
                
                # Format as mm:ss with proper rounding for accuracy
                rounded_avg_seconds = round(average_time_seconds)
                avg_minutes = int(rounded_avg_seconds // 60)
                avg_seconds = int(rounded_avg_seconds % 60)
                average_time_formatted = f"{avg_minutes}:{avg_seconds:02d}"
                
                leaderboard_data.append({
                    'student': student,
                    'average_percentage': round(average_percentage, 2),
                    'average_time_seconds': average_time_seconds,  # For sorting
                    'average_time': average_time_formatted,  # For display (mm:ss)
                    'exams_attempted': exams_attempted,
                    'exam_details': exam_details,
                })
    
    # Sort by ranking priority (consistent with class_leaderboard):
    # 1. Number of Test exams taken (more tests = higher rank)
    # 2. Average Test score percentage (higher percentage = higher rank)
    # 3. Average Test completion time (lower time = higher rank)
    leaderboard_data.sort(key=lambda x: (-x['exams_attempted'], -x['average_percentage'], x['average_time_seconds']))
    
    # Add rank with proper tie handling (considering all three criteria)
    # Students with same exams_attempted, average_percentage, AND average_time get the same rank
    for i, data in enumerate(leaderboard_data):
        if i == 0:
            data['rank'] = 1
        else:
            prev_data = leaderboard_data[i - 1]
            if (data['exams_attempted'] == prev_data['exams_attempted'] and
                data['average_percentage'] == prev_data['average_percentage'] and 
                data['average_time_seconds'] == prev_data['average_time_seconds']):
                # Same on all three criteria = same rank
                data['rank'] = prev_data['rank']
            else:
                # Different on any criteria = next sequential rank
                data['rank'] = i + 1
    
    context = {
        'page_title': 'Class Rankings',
        'enrolled_classes': enrolled_classes,
        'selected_class': selected_class,
        'leaderboard_data': leaderboard_data,
        'total_exams': assigned_exams.count(),
        'total_students': students.count(),
        'participating_students': len(leaderboard_data),
    }
    
    return render(request, 'leaderboard/student_rankings.html', context)


@login_required
def profile_settings(request):
    """
    Minimal profile settings page for both students and teachers.
    Allows users to update their basic profile information.
    """
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('profile_settings')
    else:
        form = ProfileUpdateForm(instance=request.user)
    
    context = {
        'page_title': 'Profile Settings',
        'form': form,
    }
    
    return render(request, 'settings/profile_settings.html', context)


@login_required
def teacher_change_password(request):
    """
    Teacher Password Change - Allow teachers to change their own password
    
    Security:
    - Requires current password verification
    - Validates new password requirements
    - Only accessible by teachers
    - User can only change their own password
    """
    # Verify user is a teacher
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. This feature is only available for teachers.')
        return redirect('profile_settings')
    
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validation
        if not current_password or not new_password or not confirm_password:
            messages.error(request, 'All password fields are required.')
            return redirect('teacher_change_password')
        
        # Verify current password
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('teacher_change_password')
        
        # Check new password and confirmation match
        if new_password != confirm_password:
            messages.error(request, 'New password and confirmation do not match.')
            return redirect('teacher_change_password')
        
        # Validate new password length
        if len(new_password) < 8:
            messages.error(request, 'New password must be at least 8 characters long.')
            return redirect('teacher_change_password')
        
        # Check that new password is different from current
        if current_password == new_password:
            messages.error(request, 'New password must be different from your current password.')
            return redirect('teacher_change_password')
        
        # Change the password
        request.user.set_password(new_password)
        request.user.save()
        
        # Update session to prevent logout
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)
        
        messages.success(
            request,
            'Your password has been changed successfully!'
        )
        return redirect('profile_settings')
    
    # GET request - show form
    context = {
        'page_title': 'Change Password',
    }
    
    return render(request, 'settings/change_password.html', context)


@login_required
def teacher_reset_student_password(request, student_id):
    """
    Teacher Password Reset for Students
    
    Allows a class teacher to reset a student's password, but ONLY if:
    1. The requester is a teacher
    2. The student belongs to EXACTLY ONE active class
    3. The requester is the teacher who created that specific class
    4. The target user is actually a student
    
    Security: Strict role-based + Single-class ownership validation
    """
    # Verify requester is a teacher
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Only teachers can reset student passwords.')
        return redirect('dashboard')
    
    # Get the student user
    student = get_object_or_404(User, id=student_id, role='student', is_active=True)
    
    # Get ALL active classes the student belongs to
    student_classes = StudentClass.objects.filter(
        students=student,
        is_active=True
    )
    
    # Enforce: Student must belong to EXACTLY ONE class
    if student_classes.count() == 0:
        messages.error(request, 'This student does not belong to any active class.')
        return redirect('class_list')
    
    if student_classes.count() > 1:
        messages.error(
            request, 
            'Security restriction: This student belongs to multiple classes. '
            'Password reset is only allowed for students in exactly one class.'
        )
        return redirect('class_list')
    
    # Get the ONE class the student belongs to
    student_class = student_classes.first()
    
    # Verify the requesting teacher is the creator of THIS specific class
    if student_class.created_by != request.user:
        messages.error(
            request, 
            f'Access denied. Only {student_class.created_by.get_full_name()} '
            f'(the teacher of {student_class.name}) can reset this student\'s password.'
        )
        return redirect('class_list')
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validation
        if not new_password or not confirm_password:
            messages.error(request, 'Both password fields are required.')
            return redirect('teacher_reset_student_password', student_id=student_id)
        
        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('teacher_reset_student_password', student_id=student_id)
        
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('teacher_reset_student_password', student_id=student_id)
        
        # Reset the password
        student.set_password(new_password)
        student.save()
        
        messages.success(
            request, 
            f'Password for {student.get_full_name()} (@{student.username}) has been reset successfully.'
        )
        return redirect('class_detail', class_pk=student_class.id)
    
    # GET request - show confirmation form
    context = {
        'page_title': 'Reset Student Password',
        'student': student,
        'student_class': student_class,
    }
    
    return render(request, 'teacher/reset_student_password.html', context)


@login_required
def teacher_retest_exam(request, attempt_pk):
    """
    Teacher Retest Exam - Allow Student to Retake Test Attempt
    
    Allows a class teacher to delete a student's Test attempt, enabling them to retake it.
    
    Security Requirements:
    1. Requester must be a teacher
    2. Student must belong to EXACTLY ONE active class
    3. The teacher must be the creator of that specific class
    4. The exam must be assigned to that class
    5. The exam type must be 'practice_test' or 'test'
    6. The attempt mode must be 'test' (not practice)
    7. The attempt must be completed
    
    Philosophy: Strict access control to prevent unauthorized attempt deletion
    """
    # Verify requester is a teacher
    if not request.user.is_teacher:
        messages.error(request, 'Access denied. Only teachers can authorize retests.')
        return redirect('dashboard')
    
    # Get the attempt
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related('student', 'exam', 'student_class'),
        pk=attempt_pk,
        is_completed=True
    )
    
    # Verify attempt is a Test attempt (not Practice)
    if attempt.attempt_mode != 'test':
        messages.error(request, 'Retest is only allowed for Test attempts, not Practice attempts.')
        return redirect('teacher_view_attempt', attempt_pk=attempt_pk)
    
    # Verify exam type is eligible for retest (practice_test or test)
    if attempt.exam.exam_type not in ['practice_test', 'test']:
        messages.error(request, 'This exam type does not support retests.')
        return redirect('teacher_view_attempt', attempt_pk=attempt_pk)
    
    student = attempt.student
    exam = attempt.exam
    
    # Get ALL active classes the student belongs to
    student_classes = StudentClass.objects.filter(
        students=student,
        is_active=True
    )
    
    # Enforce: Student must belong to EXACTLY ONE class
    if student_classes.count() == 0:
        messages.error(request, 'This student does not belong to any active class.')
        return redirect('teacher_exam_responses')
    
    if student_classes.count() > 1:
        messages.error(
            request,
            'Security restriction: This student belongs to multiple classes. '
            'Retest authorization is only allowed for students in exactly one class.'
        )
        return redirect('teacher_exam_responses')
    
    # Get the ONE class the student belongs to
    student_class = student_classes.first()
    
    # Verify the requesting teacher is the creator of THIS specific class
    if student_class.created_by != request.user:
        messages.error(
            request,
            f'Access denied. Only {student_class.created_by.get_full_name()} '
            f'(the teacher of {student_class.name}) can authorize retests for this student.'
        )
        return redirect('teacher_exam_responses')
    
    # Verify the exam is assigned to this class
    if not student_class.assigned_exams.filter(id=exam.id).exists():
        messages.error(
            request,
            f'This exam is not assigned to {student_class.name}. '
            'Retest authorization is only allowed for exams assigned to the student\'s class.'
        )
        return redirect('teacher_exam_responses')
    
    if request.method == 'POST':
        # Delete the Test attempt to allow retake
        student_name = student.get_full_name() or student.username
        exam_title = exam.title
        attempt_score = attempt.percentage
        
        attempt.delete()
        
        messages.success(
            request,
            f'Test attempt deleted successfully. {student_name} can now retake "{exam_title}" '
            f'(Previous score: {attempt_score:.1f}%).'
        )
        return redirect('teacher_exam_responses')
    
    # GET request - show confirmation page
    context = {
        'page_title': 'Authorize Retest',
        'attempt': attempt,
        'student': student,
        'exam': exam,
        'student_class': student_class,
    }
    
    return render(request, 'teacher/retest_confirmation.html', context)


# ============================================================================
# SCHOOL ADMIN VIEWS
# All views below require role='admin' and a school FK.
# These provide user management scoped strictly to one school.
# ============================================================================

def school_admin_required(view_func):
    """Decorator: requires school admin role (role='admin' with a school)"""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_school_admin:
            messages.error(request, 'Access denied. School administrators only.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def school_admin_dashboard(request):
    """
    School Admin Dashboard — overview of the school's data.
    Strictly scoped to request.user.school.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied. School administrators only.')
        return redirect('dashboard')

    school = request.user.school

    # Counts
    teacher_count = User.objects.filter(school=school, role='teacher', is_active=True).count()
    student_count = User.objects.filter(school=school, role='student', is_active=True).count()
    class_count = StudentClass.objects.filter(school=school, is_active=True).count()
    exam_count = Exam.objects.filter(school=school).count()
    attempt_count = ExamAttempt.objects.filter(
        student__school=school, is_completed=True
    ).count()

    # Recent teachers
    recent_teachers = User.objects.filter(
        school=school, role='teacher'
    ).order_by('-date_joined')[:5]

    # Recent students
    recent_students = User.objects.filter(
        school=school, role='student'
    ).order_by('-date_joined')[:5]

    context = {
        'school': school,
        'teacher_count': teacher_count,
        'student_count': student_count,
        'class_count': class_count,
        'exam_count': exam_count,
        'attempt_count': attempt_count,
        'recent_teachers': recent_teachers,
        'recent_students': recent_students,
        'page_title': f'{school.name} — Admin Dashboard',
    }
    return render(request, 'school_admin/dashboard.html', context)


@login_required
def school_admin_manage_teachers(request):
    """
    List all teachers in the school. School admin can deactivate/reactivate teachers.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    teachers = User.objects.filter(
        school=school, role='teacher'
    ).order_by('-date_joined')

    search = request.GET.get('q', '').strip()
    if search:
        teachers = teachers.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search)
        )

    # Implement pagination (20 items per page)
    from django.core.paginator import Paginator
    paginator = Paginator(teachers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'school': school,
        'teachers': page_obj,
        'page_obj': page_obj,
        'total_count': paginator.count,
        'search': search,
        'page_title': 'Manage Teachers',
    }
    return render(request, 'school_admin/manage_teachers.html', context)


@login_required
def school_admin_create_teacher(request):
    """
    School admin creates a teacher account for their school.
    Username is system-generated. No public registration.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school

    if request.method == 'POST':
        form = TeacherCreationForm(request.POST)
        if form.is_valid():
            teacher = form.save(school=school)
            messages.success(
                request,
                f'Teacher account created: {teacher.get_full_name()} '
                f'(Username: {teacher.username})'
            )
            return redirect('school_admin_manage_teachers')
    else:
        form = TeacherCreationForm()

    context = {
        'school': school,
        'form': form,
        'page_title': 'Create Teacher Account',
    }
    return render(request, 'school_admin/create_teacher.html', context)


@login_required
def school_admin_toggle_teacher(request, teacher_id):
    """Activate or deactivate a teacher account within the school."""
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    teacher = get_object_or_404(User, pk=teacher_id, school=school, role='teacher')

    if request.method == 'POST':
        teacher.is_active = not teacher.is_active
        teacher.save(update_fields=['is_active'])
        status = 'activated' if teacher.is_active else 'deactivated'
        messages.success(request, f'Teacher {teacher.get_full_name()} has been {status}.')

    return redirect('school_admin_manage_teachers')


@login_required
def school_admin_bulk_create_teachers(request):
    """
    School admin bulk-creates teachers from a CSV/Excel file.
    CSV columns: first_name, last_name, email (optional), phone_number (optional), password (optional)
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school

    if request.method == 'POST':
        from .forms import BulkTeacherUploadForm
        form = BulkTeacherUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            filename = uploaded_file.name.lower()

            created_count = 0
            error_rows = []

            try:
                if filename.endswith('.csv'):
                    import csv
                    import io
                    content = uploaded_file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = list(reader)
                elif filename.endswith(('.xlsx', '.xls')):
                    from openpyxl import load_workbook
                    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
                    wb.close()
                else:
                    messages.error(request, 'Unsupported file format. Use .csv, .xlsx, or .xls')
                    return redirect('school_admin_bulk_create_teachers')

                with transaction.atomic():
                    from .models import generate_username
                    for i, row in enumerate(rows, start=2):
                        first_name = row.get('first_name', '').strip()
                        last_name = row.get('last_name', '').strip()
                        email = row.get('email', '').strip() or None
                        phone_number = row.get('phone_number', '').strip() or None
                        password = row.get('password', '').strip() or 'changeme123'

                        if not first_name or not last_name:
                            error_rows.append(f'Row {i}: missing first_name or last_name')
                            continue

                        # Generate username
                        seq = school.get_next_teacher_number()
                        candidate = generate_username(school.code, 'tch', seq)
                        base = candidate
                        counter = 1
                        while User.objects.filter(username=candidate).exists():
                            candidate = f"{base}_{counter}"
                            counter += 1

                        user = User(
                            username=candidate,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            phone_number=phone_number,
                            role='teacher',
                            school=school,
                        )
                        user.set_password(password)
                        user.save()
                        created_count += 1

                if error_rows:
                    for err in error_rows:
                        messages.warning(request, err)
                
                messages.success(
                    request,
                    f'Bulk upload complete: {created_count} teacher(s) created.'
                )
                return redirect('school_admin_manage_teachers')

            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        from .forms import BulkTeacherUploadForm
        form = BulkTeacherUploadForm()

    context = {
        'school': school,
        'form': form,
        'page_title': 'Bulk Create Teachers',
    }
    return render(request, 'school_admin/bulk_create_teachers.html', context)


@login_required
def school_admin_manage_students(request):
    """
    List all students in the school with search and filter.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    students = User.objects.filter(
        school=school, role='student'
    ).select_related('school').prefetch_related('student_classes').order_by('-date_joined')

    search = request.GET.get('q', '').strip()
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(username__icontains=search)
        )

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(students, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'school': school,
        'page_obj': page_obj,
        'total_students': students.count(),
        'search': search,
        'page_title': 'Manage Students',
    }
    return render(request, 'school_admin/manage_students.html', context)


@login_required
def school_admin_create_student(request):
    """
    School admin creates a single student account.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school

    if request.method == 'POST':
        form = StudentCreationForm(request.POST)
        if form.is_valid():
            student = form.save(school=school)
            messages.success(
                request,
                f'Student account created: {student.get_full_name()} '
                f'(Username: {student.username})'
            )
            return redirect('school_admin_manage_students')
    else:
        form = StudentCreationForm()

    context = {
        'school': school,
        'form': form,
        'page_title': 'Create Student Account',
    }
    return render(request, 'school_admin/create_student.html', context)


@login_required
def school_admin_toggle_student(request, student_id):
    """Activate or deactivate a student account within the school."""
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    student = get_object_or_404(User, pk=student_id, school=school, role='student')

    if request.method == 'POST':
        student.is_active = not student.is_active
        student.save(update_fields=['is_active'])
        status = 'activated' if student.is_active else 'deactivated'
        messages.success(request, f'Student {student.get_full_name()} has been {status}.')

    return redirect('school_admin_manage_students')


@login_required
def school_admin_bulk_create_students(request):
    """
    School admin bulk-creates students from a CSV/Excel file.
    CSV columns: first_name, last_name, identifier, password (optional), email (optional)
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school

    if request.method == 'POST':
        form = BulkStudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            filename = uploaded_file.name.lower()

            created_count = 0
            error_rows = []

            try:
                if filename.endswith('.csv'):
                    import csv
                    import io
                    content = uploaded_file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = list(reader)
                elif filename.endswith(('.xlsx', '.xls')):
                    from openpyxl import load_workbook
                    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
                    wb.close()
                else:
                    messages.error(request, 'Unsupported file format. Use .csv, .xlsx, or .xls')
                    return redirect('school_admin_bulk_create_students')

                with transaction.atomic():
                    from .models import generate_username
                    import re
                    for i, row in enumerate(rows, start=2):
                        first_name = row.get('first_name', '').strip()
                        last_name = row.get('last_name', '').strip()
                        identifier = row.get('identifier', '').strip()
                        password = row.get('password', '').strip() or 'changeme123'
                        email = row.get('email', '').strip() or None

                        if not first_name or not last_name or not identifier:
                            error_rows.append(f'Row {i}: missing first_name, last_name or identifier')
                            continue

                        # Generate username
                        clean_id = re.sub(r'[^a-zA-Z0-9\-]', '', identifier)
                        candidate = generate_username(school.code, 'stu', clean_id)
                        base = candidate
                        counter = 1
                        while User.objects.filter(username=candidate).exists():
                            candidate = f"{base}_{counter}"
                            counter += 1

                        user = User(
                            username=candidate,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            role='student',
                            school=school,
                        )
                        user.set_password(password)
                        user.save()
                        created_count += 1

                if error_rows:
                    for err in error_rows:
                        messages.warning(request, err)

                messages.success(
                    request,
                    f'Bulk upload complete: {created_count} student(s) created.'
                )
                return redirect('school_admin_manage_students')

            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = BulkStudentUploadForm()

    context = {
        'school': school,
        'form': form,
        'page_title': 'Bulk Create Students',
    }
    return render(request, 'school_admin/bulk_create_students.html', context)


@login_required
def school_admin_classes_overview(request):
    """
    School admin view: overview of all classes in the school across all teachers.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    classes = StudentClass.objects.filter(
        school=school
    ).select_related('created_by').prefetch_related('students', 'assigned_exams').order_by('-year', '-created_at')

    context = {
        'school': school,
        'classes': classes,
        'total_classes': classes.count(),
        'active_classes': classes.filter(is_active=True).count(),
        'page_title': 'All Classes',
    }
    return render(request, 'school_admin/classes_overview.html', context)


@login_required
def school_admin_reset_user_password(request, user_id):
    """
    School admin resets any user's password within their school.
    """
    if not request.user.is_school_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    school = request.user.school
    target_user = get_object_or_404(User, pk=user_id, school=school)

    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if len(new_password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
        elif new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
        else:
            target_user.set_password(new_password)
            target_user.save()
            messages.success(
                request,
                f'Password reset for {target_user.get_full_name()} (@{target_user.username}).'
            )
            if target_user.role == 'teacher':
                return redirect('school_admin_manage_teachers')
            return redirect('school_admin_manage_students')

    context = {
        'school': school,
        'target_user': target_user,
        'page_title': f'Reset Password — {target_user.get_full_name()}',
    }
    return render(request, 'school_admin/reset_user_password.html', context)